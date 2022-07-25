from dataclasses import dataclass, asdict
import os
import numpy as np
import pandas as pd
import openpyxl
import csv
from pathlib import Path

from steam_sdk.data.DataLEDET import LEDETInputs, LEDETOptions, LEDETPlots, LEDETVariables, LEDETAuxiliary
from steam_sdk.data.DataRoxieParser import RoxieData
from steam_sdk.data import DataModelMagnet, DataModelConductor
from steam_sdk.data import DictionaryLEDET
from steam_sdk.configs.tools_defaults.ToolDefaultReader import ToolDefaultReader
from steam_sdk.builders.geometricFunctions import close_pairs_ckdtree, close_pairs_pdist
from steam_sdk.builders.SelfMutualInductanceCalculation import SelfMutualInductanceCalculation
from steam_sdk.parsers.ParserMap2d import getParametersFromMap2d
from steam_sdk.plotters.PlotterModel import PlotterModel


class BuilderLEDET:
    """
        Class to generate LEDET models
    """

    def __init__(self, path_magnet: Path = None, input_model_data = None,
                 input_roxie_data: RoxieData = None, input_map2d: str = None,
                 case_model: str = 'magnet',
                 flag_build: bool = True, flag_plot_all: bool = False,
                 verbose: bool = True, smic_write_path: str = ''):
        """
            Object is initialized by defining LEDET variable structure and default parameter descriptions.
            Optionally, the argument model_data can be passed to read the variables of a BuilderModel object.
            If flagInstantBuild is set to True, the LEDET input file is generated.
            If verbose is set to True, additional information will be displayed.
            case_model is a string defining the type of model to build (magnet or conductor)
        """

        # Initialize the model data structure.
        # Note: Different model types (defined by case_model) require a different data structure
        if case_model == 'magnet':
            self.model_data: DataModelMagnet = input_model_data
        elif case_model == 'conductor':
            self.model_data: DataModelConductor = input_model_data
        else:
            raise Exception('Case model {} is not supported when building a LEDET model.'.format(case_model))

        self.verbose: bool = verbose
        self.path_magnet: Path = path_magnet
        self.roxie_data: RoxieData = input_roxie_data
        self.input_map2d = input_map2d
        self.case_model = case_model
        self.flag_build = flag_build
        self.flag_plot_all = flag_plot_all

        # Data structures
        self.Inputs = LEDETInputs()
        self.Options = LEDETOptions()
        self.Plots = LEDETPlots()
        self.Variables = LEDETVariables()
        self.Auxiliary = LEDETAuxiliary()

        self.descriptionsInputs = {}
        self.descriptionsOptions = {}
        self.descriptionsPlots = {}
        self.descriptionsVariables = {}
        self.sectionTitles = {}

        # Misc
        self.smic_write_path: str = smic_write_path
        self.enableConductorResistanceFraction = False

        # Load and set the default parameter descriptions
        self.loadDefaultVariableDescriptions(os.path.join('LEDET', 'variableNamesDescriptions.xlsx'))

        if self.case_model == 'magnet':
            if (not self.model_data or not self.roxie_data) and flag_build:
                raise Exception('Cannot build model without providing DataModelMagnet and RoxieData')

            if flag_build:
                # Add method to translate all LEDET parameters from model_data to LEDET dataclasses
                self.translateModelDataToLEDET()  # TODO: fix this method (it doesn't pass the test_readFromExcel() test)

                # Find winding geometry set in the input file
                self.loadTypeWindings()
                type_windings = DictionaryLEDET.lookupWindings(self.Options.flag_typeWindings, mode='ledet2data')

                # Read geometric information
                if self.input_map2d:
                    # Read ROXIE-generated .map2d file, load some conductor parameters, and calculate other parameters
                    self.loadParametersFromMap2d(path_map2d=self.input_map2d, flag_plot=self.flag_plot_all)
                else:
                    if verbose: print('Map2d file not defined. Some geometry parameters will be read from the input yaml file.')
                    self.loadParametersFromDataModel()

                # Load conductor data from DataModelMagnet keys
                self.loadConductorData()

                # Calculate electrical order of the half-turns
                self.calcElectricalOrder(flag_plot=self.flag_plot_all)

                # Calculate thermal connections, i.e. pairs of half-turns that are in thermal contact
                if type_windings in ['multipole', 'solenoid', 'busbar']:  # notably, CCT geometry is not handled here
                    self.setThermalConnections()

                # Add thermal connections which where manually set
                self.addThermalConnections()

                # Remove thermal connections which where manually set
                self.removeThermalConnections(flag_plot=self.flag_plot_all)

                # Calculate self-mutual inductances between magnet coil sections and turns
                if self.Auxiliary.flag_calculate_inductance:
                    self.calculateSelfMutualInductance(csv_write_path=self.smic_write_path)
                else:
                    self.setSelfMutualInductances()

        elif self.case_model == 'conductor':
            if not self.model_data and flag_build:
                raise Exception('Cannot build model without providing DataModelConductor')

            if flag_build:
                # Add method to translate all LEDET parameters from model_data to LEDET dataclasses
                self.translateModelDataToLEDET()
                self.Options.flag_typeWindings = DictionaryLEDET.lookupWindings('busbar', mode='data2ledet')

                # Load conductor data from DataModelConductor keys
                self.loadConductorData(overwrite_conductor_to_group = [1])

                # Assign default values to LEDET variables defining coil windings parameters
                self.assignDefaultValuesWindings()


    def loadDefaultVariableDescriptions(self, fileVariableDescriptions: str):
        """
            **Loads and sets the LEDET descriptions**

            Function to load and set the descriptions of LEDET parameters

            :param fileName: String defining the name of the file defining the default LEDET descriptions
            :type fileName: str

            :return: None
        """
        # Read variable names and descriptions
        fullfileName = ToolDefaultReader.getResourcePath(fileVariableDescriptions)
        workbookVariables = openpyxl.load_workbook(fullfileName)

        # Load "Inputs" sheet
        worksheetInputs = workbookVariables['Inputs']
        descriptionsInputs = {}
        previousVar = 'Start_I'
        for i in range(1, worksheetInputs.max_row+1):

            if str(worksheetInputs.cell(i, 2).value)=='None' and str(worksheetInputs.cell(i, 1).value)!=None:
                if str(worksheetInputs.cell(i, 1).value)=='-': continue
                if str(worksheetInputs.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetInputs.cell(i, 1).value)
                continue
            descriptionsInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            previousVar = str(worksheetInputs.cell(i, 2).value)

        # Load "Options" sheet
        worksheetOptions = workbookVariables['Options']
        descriptionsOptions = {}
        previousVar = 'Start_O'
        for i in range(1, worksheetOptions.max_row+1):
            if str(worksheetOptions.cell(i, 2).value)=='None' and str(worksheetOptions.cell(i, 1).value)!=None:
                if str(worksheetOptions.cell(i, 1).value)=='-': continue
                if str(worksheetOptions.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetOptions.cell(i, 1).value)
                continue

            descriptionsOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)
            previousVar = str(worksheetOptions.cell(i, 2).value)

        # Load "Plots" sheet
        worksheetPlots = workbookVariables['Plots']
        descriptionsPlots = {}
        previousVar = 'Start_P'
        for i in range(1, worksheetPlots.max_row+1):
            if str(worksheetPlots.cell(i, 2).value)=='None' and str(worksheetPlots.cell(i, 1).value)!=None:
                if str(worksheetPlots.cell(i, 1).value)=='-': continue
                if str(worksheetPlots.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetPlots.cell(i, 1).value)
                continue
            descriptionsPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)
            previousVar = str(worksheetPlots.cell(i, 2).value)

        # Load "Variables" sheet
        worksheetVariables = workbookVariables['Variables']
        descriptionsVariables = {}
        previousVar = 'Start_V'
        for i in range(1, worksheetVariables.max_row+1):
            if str(worksheetVariables.cell(i, 2).value)=='None' and str(worksheetVariables.cell(i, 1).value)!=None:
                if str(worksheetVariables.cell(i, 1).value)=='-': continue
                if str(worksheetVariables.cell(i, 1).value) == 'None': continue
                self.sectionTitles[previousVar] = str(worksheetVariables.cell(i, 1).value)
                continue
            descriptionsVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)
            previousVar = str(worksheetVariables.cell(i, 2).value)

        # Set descriptions
        self.descriptionsInputs, self.descriptionsOptions, self.descriptionsPlots, self.descriptionsVariables = descriptionsInputs, descriptionsOptions, descriptionsPlots, descriptionsVariables


    def loadParametersFromMap2d(self, path_map2d: Path =None, flag_plot: bool = False):
        """
            ** Load auxiliary parameters to self.Inputs and self.Auxiliary parameters using map2d file from ROXIE **

            :param path_map2d: Input .map2d file. Note: By default, read the .map2d file defined in the yaml input file
            :type path_map2d: Path

            :return: None
        """
        # Acquire required parameters
        if path_map2d is None:
            path_map2d: Path = Path.joinpath(self.path_magnet, self.model_data.Sources.magnetic_field_fromROXIE)  # By default, read the .map2d file defined in the yaml input file
        headerLines: int = self.model_data.Options_LEDET.field_map_files.headerLines
        verbose = self.verbose

        nT, nStrands_inGroup_ROXIE, polarities_inGroup, strandToHalfTurn, strandToGroup, x_strands, y_strands, I_strands \
            = getParametersFromMap2d(map2dFile=path_map2d, headerLines=headerLines, verbose=verbose)

        indexTstop = np.cumsum(nT).tolist()
        indexTstart = [1]
        for i in range(len(nT) - 1):
            indexTstart.extend([indexTstart[i] + nT[i]])

        self.setAttribute(self.Inputs,    'nT', nT)
        self.setAttribute(self.Auxiliary, 'nStrands_inGroup_ROXIE', nStrands_inGroup_ROXIE)
        self.setAttribute(self.Inputs,    'polarities_inGroup', polarities_inGroup)
        self.setAttribute(self.Auxiliary, 'strandToHalfTurn', strandToHalfTurn)
        self.setAttribute(self.Auxiliary, 'strandToGroup', strandToGroup)
        self.setAttribute(self.Auxiliary, 'indexTstart', indexTstart)
        self.setAttribute(self.Auxiliary, 'indexTstop', indexTstop)
        self.setAttribute(self.Auxiliary, 'x_strands', x_strands)
        self.setAttribute(self.Auxiliary, 'y_strands', y_strands)
        self.setAttribute(self.Auxiliary, 'I_strands', I_strands)

        if flag_plot:
            PM = PlotterModel(self.roxie_data)
            PM.plot_conductor_numbering(self.model_data, strandToGroup, strandToHalfTurn, polarities_inGroup, x_strands, y_strands)


    def loadParametersFromDataModel(self):
        '''
            ** Load auxiliary parameters to self.Inputs parameters using input .yaml file **
            :return: None
        '''
        # Assign key values from the DataModel keys (originally defined in the input yaml file)
        nT = self.model_data.CoilWindings.n_half_turn_in_group
        self.setAttribute(self.Inputs,    'nT', nT)
        self.setAttribute(self.Inputs,    'polarities_inGroup', self.model_data.CoilWindings.polarities_in_group)


        # Calculate auxiliary variables and set their values
        indexTstop = np.cumsum(nT).tolist()
        indexTstart = [1]
        for i in range(len(nT) - 1):
            indexTstart.extend([indexTstart[i] + nT[i]])
        self.setAttribute(self.Auxiliary, 'indexTstart', indexTstart)
        self.setAttribute(self.Auxiliary, 'indexTstop', indexTstop)

        # To make sure these variables are present and not used afterwards, they are all set to NaN
        self.setAttribute(self.Auxiliary, 'nStrands_inGroup_ROXIE', np.NaN)
        self.setAttribute(self.Auxiliary, 'strandToHalfTurn', np.NaN)
        self.setAttribute(self.Auxiliary, 'strandToGroup', np.NaN)
        self.setAttribute(self.Auxiliary, 'x_strands', np.NaN)
        self.setAttribute(self.Auxiliary, 'y_strands', np.NaN)
        self.setAttribute(self.Auxiliary, 'I_strands', np.NaN)


    def setAttribute(self, LEDETclass, attribute: str, value):
        try:
            setattr(LEDETclass, attribute, value)
        except:
            setattr(getattr(self, LEDETclass), attribute, value)


    def getAttribute(self, LEDETclass, attribute):
        try:
            return getattr(LEDETclass, attribute)
        except:
            return getattr(getattr(self, LEDETclass), attribute)


    def translateModelDataToLEDET(self):
        """"
            Translates and sets parameters in self.DataModelMagnet to DataLEDET if parameter exists in LEDET
        """
        # Transform DataModelMagnet structure to dictionary with dot-separated branches
        df = pd.json_normalize(self.model_data.dict(), sep='.')
        dotSepModelData = df.to_dict(orient='records')[0]

        for keyModelData, value in dotSepModelData.items():
            keyLEDET = DictionaryLEDET.lookupModelDataToLEDET(keyModelData)
            if keyLEDET:
                if keyLEDET in self.Inputs.__annotations__:
                    self.setAttribute(self.Inputs, keyLEDET, value)
                elif keyLEDET in self.Options.__annotations__:
                    self.setAttribute(self.Options, keyLEDET, value)
                elif keyLEDET in self.Plots.__annotations__:
                    self.setAttribute(self.Plots, keyLEDET, value)
                elif keyLEDET in self.Variables.__annotations__:
                    self.setAttribute(self.Variables, keyLEDET, value)
                elif keyLEDET in self.Auxiliary.__annotations__:
                    self.setAttribute(self.Auxiliary, keyLEDET, value)
                else:
                    print('Warning: Can find {} in lookup table but not in DataLEDET'.format(keyLEDET))
                    # raise KeyError('Can find {} in lookup table but not in DataLEDET'.format(keyLEDET))


    def loadTypeWindings(self):
        '''
        Assign the integer number defining the type of windings / magnet geometry in LEDET
        '''
        self.Options.flag_typeWindings = DictionaryLEDET.lookupWindings(self.model_data.GeneralParameters.magnet_type, mode='data2ledet')


    def loadConductorData(self, overwrite_conductor_to_group: list = []):
        '''
            Load conductor data from DataModelMagnet keys
            overwrite_conductor_to_group is used in case the variable "overwrite_conductor_to_group" should be set manually rather than read from the data srtucture (this is always done, for example, when case_model='conductor')
        '''

        # Unpack variables
        if overwrite_conductor_to_group == []:
            conductor_to_group = self.model_data.CoilWindings.conductor_to_group  # Reminder: This key assigns to each group a conductor of one of the conductors
        else:
            conductor_to_group = overwrite_conductor_to_group  # use overwritten value

        # Initialize Cable variables that need to be set by this method
        self.nStrands_inGroup = np.array([])
        self.insulationType_inGroup = np.array([])
        self.internalVoidsType_inGroup = np.array([])
        self.externalVoidsType_inGroup = np.array([])
        self.wBare_inGroup = np.array([])
        self.hBare_inGroup = np.array([])
        self.wIns_inGroup = np.array([])
        self.hIns_inGroup = np.array([])
        self.Lp_s_inGroup = np.array([])
        self.R_c_inGroup = np.array([])
        self.overwrite_f_internalVoids_inGroup = np.array([])
        self.overwrite_f_externalVoids_inGroup = np.array([])
        # Initialize Strand variables that need to be set by this method
        self.SCtype_inGroup = np.array([])
        self.STtype_inGroup = np.array([])
        self.ds_inGroup = np.array([])
        self.f_SC_strand_inGroup = np.array([])
        self.RRR_Cu_inGroup = np.array([])
        self.Lp_f_inGroup = np.array([])
        self.f_ro_eff_inGroup = np.array([])
        self.df_inGroup = np.array([])
        # Initialize Jc-fit variables that need to be set by this method
        self.Tc0_NbTi_ht_inGroup = np.array([])
        self.Bc2_NbTi_ht_inGroup = np.array([])
        self.c1_Ic_NbTi_inGroup = np.array([])
        self.c2_Ic_NbTi_inGroup = np.array([])
        self.Tc0_Nb3Sn_inGroup = np.array([])
        self.Bc2_Nb3Sn_inGroup = np.array([])
        self.Jc_Nb3Sn0_inGroup = np.array([])
        self.alpha_Nb3Sn0_inGroup = np.array([])
        self.f_scaling_Jc_BSCCO2212_inGroup = np.array([])
        self.selectedFit_inGroup = np.array([])
        self.fitParameters_inGroup = np.empty((8, 0))  # Special case: This variable will be written as a matrix with 8 rows TODO find a solution for this special one - in LEDET this can be a matrix

        # For each group, load the cable, strand, and Jc-fit parameters according to their type
        for group, conductor_type in enumerate(conductor_to_group):
            if self.verbose: print('Group/Block #{}. Selected conductor: {}'.format(group + 1, conductor_type))
            self.loadCableData(conductor_type)
            self.loadStrandData(conductor_type)
            self.loadJcFitData(conductor_type)

        # TODO: Check that self.nStrands_inGroup is compatible with ROXIE map2d number of strands

        # Assign loaded Cable variables
        self.setAttribute(self.Inputs, 'nStrands_inGroup', self.nStrands_inGroup)
        self.setAttribute(self.Inputs, 'internalVoidsType_inGroup', self.internalVoidsType_inGroup)
        self.setAttribute(self.Inputs, 'externalVoidsType_inGroup', self.externalVoidsType_inGroup)
        self.setAttribute(self.Inputs, 'insulationType_inGroup', self.insulationType_inGroup)
        self.setAttribute(self.Inputs, 'wBare_inGroup', self.wBare_inGroup)
        self.setAttribute(self.Inputs, 'hBare_inGroup', self.hBare_inGroup)
        self.setAttribute(self.Inputs, 'wIns_inGroup', self.wIns_inGroup)
        self.setAttribute(self.Inputs, 'hIns_inGroup', self.hIns_inGroup)
        self.setAttribute(self.Inputs, 'Lp_s_inGroup', self.Lp_s_inGroup)
        self.setAttribute(self.Inputs, 'R_c_inGroup', self.R_c_inGroup)
        self.setAttribute(self.Inputs, 'overwrite_f_internalVoids_inGroup', self.overwrite_f_internalVoids_inGroup)
        self.setAttribute(self.Inputs, 'overwrite_f_externalVoids_inGroup', self.overwrite_f_externalVoids_inGroup)
        # Assign loaded Strand variables
        self.setAttribute(self.Inputs, 'SCtype_inGroup', self.SCtype_inGroup)
        self.setAttribute(self.Inputs, 'STtype_inGroup', self.STtype_inGroup)
        self.setAttribute(self.Inputs, 'ds_inGroup', self.ds_inGroup)
        self.setAttribute(self.Inputs, 'f_SC_strand_inGroup', self.f_SC_strand_inGroup)
        self.setAttribute(self.Inputs, 'RRR_Cu_inGroup', self.RRR_Cu_inGroup)
        self.setAttribute(self.Inputs, 'Lp_f_inGroup', self.Lp_f_inGroup)
        self.setAttribute(self.Inputs, 'f_ro_eff_inGroup', self.f_ro_eff_inGroup)
        self.setAttribute(self.Inputs, 'df_inGroup', self.df_inGroup)
        # Assign loaded Jc-fit variables
        self.setAttribute(self.Inputs, 'Tc0_NbTi_ht_inGroup', self.Tc0_NbTi_ht_inGroup)
        self.setAttribute(self.Inputs, 'Bc2_NbTi_ht_inGroup', self.Bc2_NbTi_ht_inGroup)
        self.setAttribute(self.Inputs, 'c1_Ic_NbTi_inGroup', self.c1_Ic_NbTi_inGroup)
        self.setAttribute(self.Inputs, 'c2_Ic_NbTi_inGroup', self.c2_Ic_NbTi_inGroup)
        self.setAttribute(self.Inputs, 'Tc0_Nb3Sn_inGroup', self.Tc0_Nb3Sn_inGroup)
        self.setAttribute(self.Inputs, 'Bc2_Nb3Sn_inGroup', self.Bc2_Nb3Sn_inGroup)
        self.setAttribute(self.Inputs, 'Jc_Nb3Sn0_inGroup', self.Jc_Nb3Sn0_inGroup)
        self.setAttribute(self.Inputs, 'alpha_Nb3Sn0_inGroup', self.alpha_Nb3Sn0_inGroup)
        self.setAttribute(self.Inputs, 'f_scaling_Jc_BSCCO2212_inGroup', self.f_scaling_Jc_BSCCO2212_inGroup)
        self.setAttribute(self.Inputs, 'selectedFit_inGroup', self.selectedFit_inGroup)
        self.setAttribute(self.Inputs, 'fitParameters_inGroup', self.fitParameters_inGroup)
        #TODO: CUDI3


    def loadCableData(self, conductor_id):
        '''
            Load the cable parameters for one group from DataModelMagnet keys, and append them to the respective self variables
        '''

        conductor = self.model_data.Conductors[conductor_id - 1]
        type_cable = conductor.cable.type
        if self.verbose: print('Conductor type: #{}. type_cable = {}'.format(conductor_id, type_cable))

        if type_cable == 'Rutherford':
            self.nStrands_inGroup                  = np.append(self.nStrands_inGroup                 , conductor.cable.n_strands)
            self.internalVoidsType_inGroup         = np.append(self.internalVoidsType_inGroup        , DictionaryLEDET.lookupInsulation(conductor.cable.material_inner_voids))
            self.externalVoidsType_inGroup         = np.append(self.externalVoidsType_inGroup        , DictionaryLEDET.lookupInsulation(conductor.cable.material_outer_voids))
            self.insulationType_inGroup            = np.append(self.insulationType_inGroup           , DictionaryLEDET.lookupInsulation(conductor.cable.material_insulation))
            self.wBare_inGroup                     = np.append(self.wBare_inGroup                    , conductor.cable.bare_cable_width)
            self.hBare_inGroup                     = np.append(self.hBare_inGroup                    , conductor.cable.bare_cable_height_mean)
            self.wIns_inGroup                      = np.append(self.wIns_inGroup                     , conductor.cable.th_insulation_along_width)
            self.hIns_inGroup                      = np.append(self.hIns_inGroup                     , conductor.cable.th_insulation_along_height)
            self.Lp_s_inGroup                      = np.append(self.Lp_s_inGroup                     , conductor.cable.strand_twist_pitch)
            self.R_c_inGroup                       = np.append(self.R_c_inGroup                      , conductor.cable.Rc)
            if conductor.cable.f_inner_voids != None:
                self.overwrite_f_internalVoids_inGroup = np.append(self.overwrite_f_internalVoids_inGroup, conductor.cable.f_inner_voids)
            if conductor.cable.f_outer_voids != None:
                self.overwrite_f_externalVoids_inGroup = np.append(self.overwrite_f_externalVoids_inGroup, conductor.cable.f_outer_voids)
        elif type_cable == 'Mono':
            self.nStrands_inGroup = np.append(self.nStrands_inGroup, 1)  # Note: The conductor is made of one single wire
            self.internalVoidsType_inGroup = np.append(self.internalVoidsType_inGroup, DictionaryLEDET.lookupInsulation(conductor.cable.material_inner_voids))
            self.externalVoidsType_inGroup = np.append(self.externalVoidsType_inGroup,
                                                       DictionaryLEDET.lookupInsulation(
                                                           conductor.cable.material_outer_voids))
            self.insulationType_inGroup = np.append(self.insulationType_inGroup, DictionaryLEDET.lookupInsulation(
                conductor.cable.material_insulation))
            self.wBare_inGroup = np.append(self.wBare_inGroup, conductor.cable.bare_cable_width)
            self.hBare_inGroup = np.append(self.hBare_inGroup, conductor.cable.bare_cable_height_mean)
            self.wIns_inGroup = np.append(self.wIns_inGroup, conductor.cable.th_insulation_along_width)
            self.hIns_inGroup = np.append(self.hIns_inGroup, conductor.cable.th_insulation_along_height)
            self.Lp_s_inGroup = np.append(self.Lp_s_inGroup, 0)  # Mono cables do not have strands
            self.R_c_inGroup = np.append(self.R_c_inGroup, 0)  # Mono cables do not have strands
            if conductor.cable.f_inner_voids != None:
                self.overwrite_f_internalVoids_inGroup = np.append(self.overwrite_f_internalVoids_inGroup,
                                                                   conductor.cable.f_inner_voids)
            if conductor.cable.f_outer_voids != None:
                self.overwrite_f_externalVoids_inGroup = np.append(self.overwrite_f_externalVoids_inGroup,
                                                                   conductor.cable.f_outer_voids)
        elif type_cable == 'Ribbon':
            self.nStrands_inGroup = np.append(self.nStrands_inGroup, 1)  # Note: "Strands" in ribbon-cables are connected in series, so the conductor is made of one single wire
            self.internalVoidsType_inGroup         = np.append(self.internalVoidsType_inGroup        , DictionaryLEDET.lookupInsulation(conductor.cable.material_inner_voids))
            self.externalVoidsType_inGroup         = np.append(self.externalVoidsType_inGroup        , DictionaryLEDET.lookupInsulation(conductor.cable.material_outer_voids))
            self.insulationType_inGroup            = np.append(self.insulationType_inGroup           , DictionaryLEDET.lookupInsulation(conductor.cable.material_insulation))
            self.wBare_inGroup                     = np.append(self.wBare_inGroup                    , conductor.cable.bare_cable_width)
            self.hBare_inGroup                     = np.append(self.hBare_inGroup                    , conductor.cable.bare_cable_height_mean)
            self.wIns_inGroup                      = np.append(self.wIns_inGroup                     , conductor.cable.th_insulation_along_width)
            self.hIns_inGroup                      = np.append(self.hIns_inGroup                     , conductor.cable.th_insulation_along_height)
            self.Lp_s_inGroup                      = np.append(self.Lp_s_inGroup                     , 0)  # Mono cables do not have strands
            self.R_c_inGroup                       = np.append(self.R_c_inGroup                      , 0)  # Mono cables do not have strands
            if conductor.cable.f_inner_voids != None:
                self.overwrite_f_internalVoids_inGroup = np.append(self.overwrite_f_internalVoids_inGroup, conductor.cable.f_inner_voids)
            if conductor.cable.f_outer_voids != None:
                self.overwrite_f_externalVoids_inGroup = np.append(self.overwrite_f_externalVoids_inGroup, conductor.cable.f_outer_voids)
        else:
            raise Exception('Group #{}. Selected cable type ({}) is not supported.'.format(conductor_id, type_cable))


    def loadStrandData(self, conductor_id):
        '''
            Load the strand parameters for one group from DataModelMagnet keys, and append them to the respective self variables
        '''

        conductor = self.model_data.Conductors[conductor_id - 1]
        type_strand = conductor.strand.type
        if self.verbose: print('Conductor type: #{}. type_strand = {}'.format(conductor_id, type_strand))

        if type_strand == 'Round':
            self.SCtype_inGroup      = np.append(self.SCtype_inGroup,      DictionaryLEDET.lookupSuperconductor(conductor.strand.material_superconductor))
            self.STtype_inGroup      = np.append(self.STtype_inGroup,      DictionaryLEDET.lookupStabilizer(conductor.strand.material_stabilizer))
            self.ds_inGroup          = np.append(self.ds_inGroup,          conductor.strand.diameter)
            self.f_SC_strand_inGroup = np.append(self.f_SC_strand_inGroup, 1 / (1 + conductor.strand.Cu_noCu_in_strand))  # f_SC=1/(1+Cu_noCu)
            self.RRR_Cu_inGroup      = np.append(self.RRR_Cu_inGroup,      conductor.strand.RRR)
            self.Lp_f_inGroup        = np.append(self.Lp_f_inGroup,        conductor.strand.fil_twist_pitch)
            self.f_ro_eff_inGroup    = np.append(self.f_ro_eff_inGroup,    conductor.strand.f_Rho_effective)
            self.df_inGroup          = np.append(self.df_inGroup,          conductor.strand.filament_diameter)
        elif type_strand == 'Rectangular':
            self.SCtype_inGroup      = np.append(self.SCtype_inGroup,      DictionaryLEDET.lookupSuperconductor(conductor.strand.material_superconductor))
            self.STtype_inGroup      = np.append(self.STtype_inGroup,      DictionaryLEDET.lookupStabilizer(conductor.strand.material_stabilizer))
            ds_equivalent            = np.sqrt( (conductor.strand.bare_width * conductor.strand.bare_height) *4/np.pi )
            self.ds_inGroup          = np.append(self.ds_inGroup,          ds_equivalent)
            self.f_SC_strand_inGroup = np.append(self.f_SC_strand_inGroup, 1 / (1 + conductor.strand.Cu_noCu_in_strand))  # f_SC=1/(1+Cu_noCu)
            self.RRR_Cu_inGroup      = np.append(self.RRR_Cu_inGroup,      conductor.strand.RRR)
            self.Lp_f_inGroup        = np.append(self.Lp_f_inGroup,        conductor.strand.fil_twist_pitch)
            self.f_ro_eff_inGroup    = np.append(self.f_ro_eff_inGroup,    conductor.strand.f_Rho_effective)
            self.df_inGroup          = np.append(self.df_inGroup,          conductor.strand.filament_diameter)
        else:
            raise Exception('Group #{}. Selected strand type ({}) is not supported.'.format(conductor_id, type_strand))


    def loadJcFitData(self, conductor_id):
        '''
            Load the Jc-fit parameters for one group from DataModelMagnet keys, and append them to the respective self variables
        '''

        conductor = self.model_data.Conductors[conductor_id - 1]
        type_JcFit = conductor.Jc_fit.type
        if self.verbose: print('Conductor type: #{}. type_JcFit = {}'.format(conductor_id, type_JcFit))

        # TODO: ConstantJc to enable for quench protection simulations
        if type_JcFit == 'ConstantJc':
            self.selectedFit_inGroup = np.append(self.selectedFit_inGroup, 1)
            temp_fitParam = np.array([
                conductor.Jc_fit.Jc_constant, 0, 0, 0, 0, 0, 0, 0])  # Last values not needed for this fit and set to 0
            self.fitParameters_inGroup = np.column_stack((self.fitParameters_inGroup, temp_fitParam))  # At each step, add one more column to the matrix
            raise Exception('Group #{}. Selected Jc-fit type ({}) is currently not supported by LEDET.'.format(conductor_id, type_JcFit))
        elif type_JcFit == 'CUDI1':
            self.Tc0_NbTi_ht_inGroup = np.append(self.Tc0_NbTi_ht_inGroup, conductor.Jc_fit.Tc0_CUDI1)
            self.Bc2_NbTi_ht_inGroup = np.append(self.Bc2_NbTi_ht_inGroup, conductor.Jc_fit.Bc20_CUDI1)
            self.c1_Ic_NbTi_inGroup  = np.append(self.c1_Ic_NbTi_inGroup  , conductor.Jc_fit.C1_CUDI1)
            self.c2_Ic_NbTi_inGroup  = np.append(self.c2_Ic_NbTi_inGroup  , conductor.Jc_fit.C2_CUDI1)
            self.Tc0_Nb3Sn_inGroup   = np.append(self.Tc0_Nb3Sn_inGroup, 0)
            self.Bc2_Nb3Sn_inGroup   = np.append(self.Bc2_Nb3Sn_inGroup, 0)
            self.Jc_Nb3Sn0_inGroup   = np.append(self.Jc_Nb3Sn0_inGroup, 0)
            self.selectedFit_inGroup = np.append(self.selectedFit_inGroup, 6)  # TODO: not yet supported by LEDET
            temp_fitParam = np.array(
                [conductor.Jc_fit.Tc0_CUDI1, conductor.Jc_fit.Bc20_CUDI1, conductor.Jc_fit.C1_CUDI1, conductor.Jc_fit.C2_CUDI1, 0, 0, 0, 0])  # Last values not needed for this fit and set to 0
            self.fitParameters_inGroup = np.column_stack((self.fitParameters_inGroup, temp_fitParam))  # At each step, add one more column to the matrix
        # TODO: CUDI3 to enable for quench protection simulations
        elif type_JcFit == 'CUDI3':
            self.selectedFit_inGroup = np.append(self.selectedFit_inGroup, 3)  # TODO: not yet supported by LEDET
            temp_fitParam = np.array([
                conductor.Jc_fit.Tc0_CUDI3, conductor.Jc_fit.Bc20_CUDI3, conductor.Jc_fit.c1_CUDI3,
                conductor.Jc_fit.c2_CUDI3, conductor.Jc_fit.c3_CUDI3, conductor.Jc_fit.c4_CUDI3,
                conductor.Jc_fit.c5_CUDI3, conductor.Jc_fit.c6_CUDI3])
            self.fitParameters_inGroup = np.column_stack((self.fitParameters_inGroup, temp_fitParam))  # At each step, add one more column to the matrix
            raise Exception('Group #{}. Selected Jc-fit type ({}) is currently not supported by LEDET.'.format(conductor_id, type_JcFit))
        # TODO: Bottura to enable for quench protection simulations
        elif type_JcFit == 'Bottura':
            self.selectedFit_inGroup = np.append(self.selectedFit_inGroup, 2)  # TODO: not yet supported by LEDET
            temp_fitParam = np.array([
                conductor.Jc_fit.Tc0_Bottura, conductor.Jc_fit.Bc20_Bottura, conductor.Jc_fit.Jc_ref_Bottura,
                conductor.Jc_fit.C0_Bottura, conductor.Jc_fit.alpha_Bottura, conductor.Jc_fit.beta_Bottura,
                conductor.Jc_fit.gamma_Bottura, 0])  # Last values not needed for this fit and set to 0
            self.fitParameters_inGroup = np.column_stack((self.fitParameters_inGroup, temp_fitParam))  # At each step, add one more column to the matrix
            raise Exception('Group #{}. Selected Jc-fit type ({}) is currently not supported by LEDET.'.format(conductor_id, type_JcFit))
        elif type_JcFit == 'Summers':
            self.Tc0_NbTi_ht_inGroup            = np.append(self.Tc0_NbTi_ht_inGroup  , 0)
            self.Bc2_NbTi_ht_inGroup            = np.append(self.Bc2_NbTi_ht_inGroup  , 0)
            self.c1_Ic_NbTi_inGroup             = np.append(self.c1_Ic_NbTi_inGroup   , 0)
            self.c2_Ic_NbTi_inGroup             = np.append(self.c2_Ic_NbTi_inGroup   , 0)
            self.Tc0_Nb3Sn_inGroup              = np.append(self.Tc0_Nb3Sn_inGroup    , conductor.Jc_fit.Tc0_Summers)
            self.Bc2_Nb3Sn_inGroup              = np.append(self.Bc2_Nb3Sn_inGroup    , conductor.Jc_fit.Bc20_Summers)
            self.Jc_Nb3Sn0_inGroup              = np.append(self.Jc_Nb3Sn0_inGroup    , conductor.Jc_fit.Jc0_Summers)
            self.selectedFit_inGroup = np.append(self.selectedFit_inGroup             , 4)
            temp_fitParam = np.array([conductor.Jc_fit.Tc0_Summers, conductor.Jc_fit.Bc20_Summers, conductor.Jc_fit.Jc0_Summers, 0, 0, 0, 0, 0])  # Last values not needed for this fit and set to 0
            self.fitParameters_inGroup          = np.column_stack((self.fitParameters_inGroup, temp_fitParam))  # At each step, add one more column to the matrix
        elif type_JcFit == 'Bordini':
            self.Tc0_NbTi_ht_inGroup            = np.append(self.Tc0_NbTi_ht_inGroup  , 0)
            self.Bc2_NbTi_ht_inGroup            = np.append(self.Bc2_NbTi_ht_inGroup  , 0)
            self.c1_Ic_NbTi_inGroup             = np.append(self.c1_Ic_NbTi_inGroup   , 0)
            self.c2_Ic_NbTi_inGroup             = np.append(self.c2_Ic_NbTi_inGroup   , 0)
            self.Tc0_Nb3Sn_inGroup              = np.append(self.Tc0_Nb3Sn_inGroup    , conductor.Jc_fit.Tc0_Bordini)
            self.Bc2_Nb3Sn_inGroup              = np.append(self.Bc2_Nb3Sn_inGroup    , conductor.Jc_fit.Bc20_Bordini)
            self.Jc_Nb3Sn0_inGroup              = np.append(self.Jc_Nb3Sn0_inGroup    , conductor.Jc_fit.C0_Bordini)
            self.alpha_Nb3Sn0_inGroup           = np.append(self.alpha_Nb3Sn0_inGroup , conductor.Jc_fit.alpha_Bordini)
            self.selectedFit_inGroup            = np.append(self.selectedFit_inGroup  , 5)  # TODO: not yet supported by LEDET
            temp_fitParam = np.array([conductor.Jc_fit.Tc0_Bordini, conductor.Jc_fit.Bc20_Bordini, conductor.Jc_fit.C0_Bordini, conductor.Jc_fit.alpha_Bordini, 0, 0, 0, 0])  # Last values not needed for this fit and set to 0
            self.fitParameters_inGroup          = np.column_stack((self.fitParameters_inGroup, temp_fitParam))  # At each step, add one more column to the matrix
        elif type_JcFit == 'BSCCO_2212_LBNL':
            self.Tc0_NbTi_ht_inGroup            = np.append(self.Tc0_NbTi_ht_inGroup  , 0)
            self.Bc2_NbTi_ht_inGroup            = np.append(self.Bc2_NbTi_ht_inGroup  , 0)
            self.c1_Ic_NbTi_inGroup             = np.append(self.c1_Ic_NbTi_inGroup   , 0)
            self.c2_Ic_NbTi_inGroup             = np.append(self.c2_Ic_NbTi_inGroup   , 0)
            self.Tc0_Nb3Sn_inGroup              = np.append(self.Tc0_Nb3Sn_inGroup    , 0)
            self.Bc2_Nb3Sn_inGroup              = np.append(self.Bc2_Nb3Sn_inGroup    , 0)
            self.Jc_Nb3Sn0_inGroup              = np.append(self.Jc_Nb3Sn0_inGroup    , 0)
            self.f_scaling_Jc_BSCCO2212_inGroup = np.append(self.f_scaling_Jc_BSCCO2212_inGroup, conductor.Jc_fit.f_scaling_Jc_BSCCO2212)
            self.selectedFit_inGroup            = np.append(self.selectedFit_inGroup  , 51)  # TODO: not yet supported by LEDET
            temp_fitParam = np.array([conductor.Jc_fit.f_scaling_Jc_BSCCO2212, 0, 0, 0, 0, 0, 0, 0])  # Last values not needed for this fit and set to 0
            self.fitParameters_inGroup          = np.column_stack((self.fitParameters_inGroup, temp_fitParam))  # At each step, add one more column to the matrix
        else:
            raise Exception('Group #{}. Selected Jc-fit type ({}) is not supported.'.format(conductor_id, type_JcFit))


    def calcElectricalOrder(self, flag_plot: bool = False):
        """
            **Calculates electrical order of each half turn (for multipole magnets) or of each turn (for solenoid and CCT magnets) **

            :param elPairs_GroupTogether: list of 2-element tuples that includes the electrical order of groups
                (e.g. [[1,17], [2,18]] means group 1 and 17 are electrically connected)
            :param elPairs_RevElOrder: list defining whether the (half-)turns present in the groups need to be electrically ordered following their order in the group (=0), or reversed (=1)
        """
        # Unpack variables
        elPairs_GroupTogether      = self.Auxiliary.elPairs_GroupTogether
        elPairs_RevElOrder         = self.Auxiliary.elPairs_RevElOrder
        nT                         = self.Inputs.nT
        indexTstart                = self.Auxiliary.indexTstart
        indexTstop                 = self.Auxiliary.indexTstop
        flag_typeWindings          = self.Options.flag_typeWindings
        typeWindings = DictionaryLEDET.lookupWindings(flag_typeWindings, mode='ledet2data')

        if self.verbose:
            print('Setting the electrical order')

        # If the key overwrite_electrical_order is defined, the electrical order is
        if self.model_data and len(self.model_data.CoilWindings.electrical_pairs.overwrite_electrical_order) > 0:
            if self.verbose: print('Electrical order is defined manually based on the input key CoilWindings.electrical_pairs.overwrite_electrical_order')

            el_order_half_turns = self.model_data.CoilWindings.electrical_pairs.overwrite_electrical_order
            # Assign values to the attribute in the LEDET Inputs dataclass
            self.setAttribute(self.Inputs, 'el_order_half_turns', np.array(el_order_half_turns))
            return el_order_half_turns  # stop here the function without calculating the electrical order

        el_order_half_turns = []
        if typeWindings in ['multipole', 'busbar']:
            if len(elPairs_RevElOrder) != len(elPairs_GroupTogether):
                raise Exception('Length of the vector elPairs_RevElOrder ({}) must be equal to nElPairs={}.'.format(
                    len(elPairs_RevElOrder), len(elPairs_GroupTogether)))
            for p in range(len(elPairs_GroupTogether)):
                if nT[elPairs_GroupTogether[p][0] - 1] != nT[elPairs_GroupTogether[p][1] - 1]:
                    raise Exception('Pair of groups defined by the variable elPairs_GroupTogether must have the same number of half-turns.')
                for k in range(nT[elPairs_GroupTogether[p][0] - 1]):
                    if elPairs_RevElOrder[p] == 0:
                        el_order_half_turns.append(indexTstart[elPairs_GroupTogether[p][0] - 1] + k)
                        el_order_half_turns.append(indexTstart[elPairs_GroupTogether[p][1] - 1] + k)
                    if elPairs_RevElOrder[p] == 1:
                        el_order_half_turns.append(indexTstop[elPairs_GroupTogether[p][0] - 1] - k)
                        el_order_half_turns.append(indexTstop[elPairs_GroupTogether[p][1] - 1] - k)
        elif typeWindings in ['solenoid']:
            nGroupsDefined = len(nT)
            winding_order_groups = (nGroupsDefined * [0, 1])[:nGroupsDefined]
            for p in range(nGroupsDefined, 0, -1):
                for k in range(nT[p - 1]):
                    if winding_order_groups[p - 1] == 0:
                        el_order_half_turns.append(indexTstart[p - 1] + k)
                    if winding_order_groups[p - 1] == 1:
                        el_order_half_turns.append(indexTstop[p - 1] - k)
        elif typeWindings in ['CCT']:
            nTurns = sum(nT)
            el_order_half_turns = [x for x in range(1, nTurns + 1)]

        # Assign values to the attribute in the LEDET Inputs dataclass
        self.setAttribute(self.Inputs, 'el_order_half_turns', np.array(el_order_half_turns))

        if self.verbose:
            print('Setting electrical order was successful.')

        if flag_plot:
            PM = PlotterModel(self.roxie_data)
            PM.plot_electrical_order(el_order_half_turns, elPairs_GroupTogether, self.Auxiliary.strandToGroup, self.Auxiliary.x_strands, self.Auxiliary.y_strands, self.Auxiliary.strandToHalfTurn, self.model_data)
        return np.array(el_order_half_turns)


    def setThermalConnections(self):
        # Unpack variables
        nT               = self.Inputs.nT
        max_distance     = self.Auxiliary.heat_exchange_max_distance
        indexTstart      = self.Auxiliary.indexTstart
        indexTstop       = self.Auxiliary.indexTstop
        strandToHalfTurn = self.Auxiliary.strandToHalfTurn
        strandToGroup    = self.Auxiliary.strandToGroup
        x_strands        = self.Auxiliary.x_strands
        y_strands        = self.Auxiliary.y_strands
        nGroups = int(len(nT))

        if self.verbose:
            print('Setting thermal connections')

        iContactAlongWidth_From = []
        iContactAlongWidth_To = []

        for g in range(nGroups):
            iContactAlongWidth_From.extend(range(indexTstart[g], indexTstop[g]))
            iContactAlongWidth_To.extend(range(indexTstart[g] + 1, indexTstop[g] + 1))

        if len(iContactAlongWidth_From) < 1:
            iContactAlongWidth_From.append(1)
            iContactAlongWidth_To.append(1)

        self.setAttribute(self.Inputs, 'iContactAlongWidth_From', np.array(iContactAlongWidth_From))
        self.setAttribute(self.Inputs, 'iContactAlongWidth_To',   np.array(iContactAlongWidth_To))

        if self.verbose:
            print('Heat exchange along the cable wide side - Calculated indices:')
            print('iContactAlongWidth_From = ')
            print(iContactAlongWidth_From)
            print('iContactAlongWidth_To = ')
            print(iContactAlongWidth_To)

        # Prepare input for the function close_pairs_ckdtree
        X = np.column_stack((x_strands, y_strands))

        # find all pairs of strands closer than a distance of max_d
        pairs_close = close_pairs_ckdtree(X, max_distance)

        # find pairs that belong to half-turns located in different groups
        contact_pairs = set([])
        for p in pairs_close:
            if not strandToGroup[p[0]] == strandToGroup[p[1]]:
                contact_pairs.add((strandToHalfTurn[p[0]], strandToHalfTurn[p[1]]))

        # assign the pair values to two distinct vectors
        iContactAlongHeight_From = []
        iContactAlongHeight_To = []
        for p in contact_pairs:
            iContactAlongHeight_From.append(p[0])
            iContactAlongHeight_To.append(p[1])
        # Keep arrays Non-empty
        if len(iContactAlongHeight_From) < 1:
            iContactAlongHeight_From.append(1)
            iContactAlongHeight_To.append(1)

        # find indices to order the vector iContactAlongHeight_From
        idxSort = [i for (v, i) in sorted((v, i) for (i, v) in enumerate(iContactAlongHeight_From))]

        # reorder both iContactAlongHeight_From and iContactAlongHeight_To using the indices
        iContactAlongHeight_From = [iContactAlongHeight_From[i] for i in idxSort]
        iContactAlongHeight_To   = [iContactAlongHeight_To[i] for i in idxSort]

        self.setAttribute(self.Inputs, 'iContactAlongHeight_From', np.array(iContactAlongHeight_From))
        self.setAttribute(self.Inputs, 'iContactAlongHeight_To',   np.array(iContactAlongHeight_To))

        if self.verbose:
            print('Thermal links are set.')
            print('Heat exchange along the cable narrow side - Calculated indices:')
            print('iContactAlongHeight_From = ')
            print(iContactAlongHeight_From)
            print('iContactAlongWidth_To = ')
            print(iContactAlongHeight_To)


    def addThermalConnections(self):
        '''
            **Adding manually set thermal connections to the ones which where automatically calculated before**
        '''
        # Unpack variables
        pairs_to_add_along_height = self.Auxiliary.iContactAlongHeight_pairs_to_add
        pairs_to_add_along_width  = self.Auxiliary.iContactAlongWidth_pairs_to_add
        iContactAlongHeight_From  = self.Inputs.iContactAlongHeight_From
        iContactAlongHeight_To    = self.Inputs.iContactAlongHeight_To
        iContactAlongWidth_From   = self.Inputs.iContactAlongWidth_From
        iContactAlongWidth_To     = self.Inputs.iContactAlongWidth_To

        # check for pair repetition
        pairs_to_add_along_height = list(set(map(tuple, pairs_to_add_along_height)))
        pairs_to_add_along_width = list(set(map(tuple, pairs_to_add_along_width)))
        pairs_to_add_along_height.sort()
        pairs_to_add_along_width.sort()

        # Splitting pairs in two lists
        iContactAlongHeight_From_to_add = []
        iContactAlongHeight_To_to_add = []
        if len(pairs_to_add_along_height) != 0:
            for p in pairs_to_add_along_height:
                iContactAlongHeight_From_to_add.append(p[0])
                iContactAlongHeight_To_to_add.append(p[1])
        iContactAlongWidth_From_to_add = []
        iContactAlongWidth_To_to_add = []
        if len(pairs_to_add_along_width) != 0:
            for p in pairs_to_add_along_width:
                iContactAlongWidth_From_to_add.append(p[0])
                iContactAlongWidth_To_to_add.append(p[1])

        # Appending manually set thermal connections
        iContactAlongHeight_From = np.append(iContactAlongHeight_From, iContactAlongHeight_From_to_add)
        iContactAlongHeight_To = np.append(iContactAlongHeight_To, iContactAlongHeight_To_to_add)
        iContactAlongWidth_From = np.append(iContactAlongWidth_From, iContactAlongWidth_From_to_add)
        iContactAlongWidth_To = np.append(iContactAlongWidth_To, iContactAlongWidth_To_to_add)


        # Reorder both sets of indices
        idxSort = [i for (v, i) in sorted((v, i) for (i, v) in enumerate(iContactAlongWidth_From))]
        # reorder both iContactAlongWidth_From and iContactAlongHeight_To using the indices
        iContactAlongWidth_From = [iContactAlongWidth_From[i] for i in idxSort]
        iContactAlongWidth_To   = [iContactAlongWidth_To[i] for i in idxSort]

        idxSort = [i for (v, i) in sorted((v, i) for (i, v) in enumerate(iContactAlongHeight_From))]
        # reorder both iContactAlongHeight_From and iContactAlongHeight_To using the indices
        iContactAlongHeight_From = [iContactAlongHeight_From[i] for i in idxSort]
        iContactAlongHeight_To   = [iContactAlongHeight_To[i] for i in idxSort]


        # Assign variables
        self.setAttribute(self.Inputs, 'iContactAlongHeight_From', np.array(iContactAlongHeight_From))
        self.setAttribute(self.Inputs, 'iContactAlongHeight_To', np.array(iContactAlongHeight_To))
        self.setAttribute(self.Inputs, 'iContactAlongWidth_From', np.array(iContactAlongWidth_From))
        self.setAttribute(self.Inputs, 'iContactAlongWidth_To', np.array(iContactAlongWidth_To))

        if len(pairs_to_add_along_height) != 0:
            print('Selected thermal links are added.')


    def removeThermalConnections(self, flag_plot: bool = False):
        '''
            *Removing manually set thermal connections from the ones which where automatically calculated before*
        '''
        # Unpack variables
        pairs_to_remove_along_height = self.Auxiliary.iContactAlongHeight_pairs_to_remove
        pairs_to_remove_along_width  = self.Auxiliary.iContactAlongWidth_pairs_to_remove
        iContactAlongHeight_From     = self.Inputs.iContactAlongHeight_From
        iContactAlongHeight_To       = self.Inputs.iContactAlongHeight_To
        iContactAlongWidth_From      = self.Inputs.iContactAlongWidth_From
        iContactAlongWidth_To        = self.Inputs.iContactAlongWidth_To

        # check for pair repetition
        pairs_to_remove_along_height = list(set(map(tuple, pairs_to_remove_along_height)))
        pairs_to_remove_along_width = list(set(map(tuple, pairs_to_remove_along_width)))
        pairs_to_remove_along_height.sort()
        pairs_to_remove_along_width.sort()

        # Splitting pairs in two lists
        iContactAlongHeight_From_to_remove = []
        iContactAlongHeight_To_to_remove = []
        if len(pairs_to_remove_along_height) != 0:
            for p in pairs_to_remove_along_height:
                iContactAlongHeight_From_to_remove.append(p[0])
                iContactAlongHeight_To_to_remove.append(p[1])
        iContactAlongWidth_From_to_remove = []
        iContactAlongWidth_To_to_remove = []
        if len(pairs_to_remove_along_width) != 0:
            for p in pairs_to_remove_along_width:
                iContactAlongWidth_From_to_remove.append(p[0])
                iContactAlongWidth_To_to_remove.append(p[1])

        # removing manually set thermal connections
        for i in range(len(pairs_to_remove_along_height)):
            for j in range(len(iContactAlongHeight_From)):
                if iContactAlongHeight_From[j-1] == pairs_to_remove_along_height[i][0] and iContactAlongHeight_To[j-1] == pairs_to_remove_along_height[i][1]:
                    iContactAlongHeight_From = np.delete(iContactAlongHeight_From, j-1)
                    iContactAlongHeight_To = np.delete(iContactAlongHeight_To, j-1)
        for i in range(len(pairs_to_remove_along_width)):
            for j in range(len(iContactAlongWidth_From)):
                if iContactAlongWidth_From[j-1] == pairs_to_remove_along_width[i][0] and iContactAlongWidth_To[j-1] == pairs_to_remove_along_width[i][1]:
                    iContactAlongWidth_From = np.delete(iContactAlongWidth_From, j-1)
                    iContactAlongWidth_To = np.delete(iContactAlongWidth_To, j-1)

        # Assign variables
        self.setAttribute(self.Inputs, 'iContactAlongHeight_From', np.array(iContactAlongHeight_From))
        self.setAttribute(self.Inputs, 'iContactAlongHeight_To', np.array(iContactAlongHeight_To))
        self.setAttribute(self.Inputs, 'iContactAlongWidth_From', np.array(iContactAlongWidth_From))
        self.setAttribute(self.Inputs, 'iContactAlongWidth_To', np.array(iContactAlongWidth_To))


        if len(pairs_to_remove_along_height) != 0:
            print('Selected thermal links are removed.')

        if flag_plot:
            PM = PlotterModel(self.roxie_data)
            PM.plot_heat_connections(iContactAlongHeight_From, iContactAlongHeight_To, iContactAlongWidth_From, iContactAlongWidth_To, self.Auxiliary.x_strands, self.Auxiliary.y_strands, self.Auxiliary.strandToHalfTurn, self.model_data)


    def calculateSelfMutualInductance(self, csv_write_path: str = ''):
        """
            Calculate the self-mutual inductance matrix
            Calculation based on the SMIC code (https://cernbox.cern.ch/index.php/s/37F87v3oeI2Gkp3)
        """
        # TODO: since many parameters, maybe access as .self when used instead of unpacking
        # TODO: however, should be stated, e.g. in docstring, which parameters needed from the outside in order to run
        name_magnet        = self.model_data.GeneralParameters.magnet_name
        nT                 = self.Inputs.nT
        nStrands_inGroup   = self.Auxiliary.nStrands_inGroup_ROXIE  # Note: This vector of number of strands in each group/block is read from the ROXIE map2d file, and it might differ from self.Inputs.nStrands_inGroup in case of Rutherford cables with an odd number of strands
        ds_inGroup         = self.Inputs.ds_inGroup
        hBare_inGroup      = self.Inputs.hBare_inGroup
        GroupToCoilSection = self.Inputs.GroupToCoilSection
        polarities_inGroup = self.Inputs.polarities_inGroup  # from map2d input file
        polarities_inGroup_from_yaml = self.model_data.CoilWindings.polarities_in_group  # from yaml input file
        if not polarities_inGroup == polarities_inGroup_from_yaml:
            raise Exception ('polarities_inGroup differs from polarities_inGroup_from_yaml')
        indexTstart        = self.Auxiliary.indexTstart
        indexTstop         = self.Auxiliary.indexTstop
        x_strands          = self.Auxiliary.x_strands
        y_strands          = self.Auxiliary.y_strands
        strandToHalfTurn   = self.Auxiliary.strandToHalfTurn
        nGroups, nHalfTurns, nStrands  = len(GroupToCoilSection), max(strandToHalfTurn), len(x_strands)

        # Set options
        flag_strandCorrection = 0
        flag_sumTurnToTurn = 1
        flag_writeOutput = 0

        # Calculate group to which each half-turn belongs
        HalfTurnToGroup = np.zeros((1, nHalfTurns), dtype=int)
        HalfTurnToGroup = HalfTurnToGroup[0]
        HalfTurnToCoilSection = np.zeros((1, nHalfTurns), dtype=int)
        HalfTurnToCoilSection = HalfTurnToCoilSection[0]
        for g in range(1, nGroups + 1):
            HalfTurnToGroup[indexTstart[g - 1] - 1:indexTstop[g - 1]] = g
            HalfTurnToCoilSection[indexTstart[g - 1] - 1:indexTstop[g - 1]] = GroupToCoilSection[g - 1]

        # Calculate group to which each strand belongs
        nS = np.repeat(nStrands_inGroup, nT)
        indexSstart = np.hstack([1, 1 + np.cumsum(nS[:-1])]).astype(int)
        indexSstop = np.cumsum(nS).astype(int)
        strandToGroup = np.zeros((1, nStrands), dtype=int)
        strandToGroup = strandToGroup[0]
        strandToCoilSection = np.zeros((1, nStrands), dtype=int)
        strandToCoilSection = strandToCoilSection[0]
        for ht in range(1, nHalfTurns + 1):
            strandToGroup[indexSstart[ht - 1] - 1:indexSstop[ht - 1]] = HalfTurnToGroup[ht - 1]
            strandToCoilSection[indexSstart[ht - 1] - 1:indexSstop[ht - 1]] = HalfTurnToCoilSection[ht - 1]

        polarities = np.repeat(polarities_inGroup, nT)
        polarities = np.repeat(polarities, nS.astype(int))
        for i in range(2):
            # Calculate diameter of each strand
            Ds = np.zeros((1, nStrands), dtype=float)
            Ds = Ds[0]
            for g in range(1, nGroups + 1):
                if i == 0: Ds[np.where(strandToGroup == g)] = ds_inGroup[g - 1]
                if i == 1: Ds[np.where(strandToGroup == g)] = hBare_inGroup[g - 1]

            # Define self-mutual inductance calculation object
            coil = SelfMutualInductanceCalculation(x_strands, y_strands, polarities,
                                                   nS, Ds, strandToHalfTurn, strandToCoilSection,
                                                   flag_strandCorrection, flag_sumTurnToTurn, flag_writeOutput,
                                                   name_magnet, verbose=self.verbose)

            # Calculate self-mutual inductance between half-turns, turns, and coil-sections, per unit length [H/m]
            M_halfTurns, M_turns, M_coilSections, L_magnet = \
                coil.calculateInductance(x_strands, y_strands, polarities,
                                         nS, Ds, strandToHalfTurn, strandToCoilSection,
                                         flag_strandCorrection=0)

            L_turns = M_turns
            L_turns_diag = np.diagonal(L_turns)
            L_turns_diag_rep = np.tile(L_turns_diag, (len(L_turns), 1))  # this replicates the effect of L_xx[i][i]
            denom_turns = np.sqrt(L_turns_diag_rep.T * L_turns_diag_rep)
            k_turns = L_turns / denom_turns  # matrix alt to k_turns[i][j]=L_turns[i][j]/np.sqrt(L_turns[j][j]*L_turns[i][i])

            # Check that the coupling factors are all higher than 1
            if len(k_turns[np.where(k_turns > 1)]) == 0:
                break
            else:
                assert max(nStrands_inGroup) == 1, 'Wires are not single stranded but mutual coupling factors k>1'
                print('Mutual coupling factors of some turns is k>1, re-calculate with hBare.')

        # Self-mutual inductances between coil sections, per unit length [H/m]
        self.setAttribute(self.Inputs, 'M_m', M_coilSections)

        # Self-mutual inductances between turns, per unit length [H/m]
        self.setAttribute(self.Inputs, 'M_InductanceBlock_m', M_turns)

        # Total magnet self-mutual inductance, per unit length [H/m]
        # L_magnet

        # Defining to which inductive block each half-turn belongs
        HalfTurnToInductanceBlock = np.concatenate((np.linspace(1, int(nHalfTurns / 2), int(nHalfTurns/2)),
                                                       np.linspace(1, int(nHalfTurns / 2), int(nHalfTurns/2))))
        self.setAttribute(self.Inputs, 'HalfTurnToInductanceBlock', HalfTurnToInductanceBlock)

        if self.verbose:
            print('')
            print('Total magnet self-inductance per unit length: ' + str(L_magnet) + ' H/m')

        # Write a csv file with the self-mutual inductance per unit length between each turn
        if not csv_write_path:
            csv_write_path = os.path.join(name_magnet + '_selfMutualInductanceMatrix.csv')
        with open(csv_write_path, 'w', newline='') as file:
            reader = csv.writer(file)
            reader.writerow(["Self- and mutual inductances per unit length between each turn [H/m]"])
            for i in range(M_turns.shape[0]):
                reader.writerow(M_turns[i])

        # If self-mutual inductance matrix too large, set M_m to 0 and LEDET will read from the .csv file instead
        if M_turns.shape[0] >= 201:
            M_InductanceBlock_m = np.array([0])
            self.setAttribute(self.Inputs, 'M_InductanceBlock_m', M_InductanceBlock_m)  # TODO: change test_BuilderModel to reflect this change


    def setSelfMutualInductances(self):
        '''
        Assign self-mutual inductance parameters (this is used when inductance calculation is not enabled
        :return:
        '''

        # Defining to which inductive block each half-turn belongs
        nHalfTurns = sum(self.Inputs.nT)
        magnet_type = self.model_data.GeneralParameters.magnet_type
        if self.Auxiliary.overwrite_HalfTurnToInductanceBlock:
            HalfTurnToInductanceBlock = self.Auxiliary.overwrite_HalfTurnToInductanceBlock
        else:
            if magnet_type == 'multipole':
                HalfTurnToInductanceBlock = np.concatenate((np.linspace(1, int(nHalfTurns / 2), int(nHalfTurns / 2)), np.linspace(1, int(nHalfTurns / 2), int(nHalfTurns / 2))))
            elif (magnet_type == 'solenoid') or (magnet_type == 'CCT') or (magnet_type == 'busbar'):
                HalfTurnToInductanceBlock = np.linspace(1, int(nHalfTurns), int(nHalfTurns))
            else:
                raise Exception(f'Magnet type not recognized: {magnet_type}.')

        # Self-mutual inductances between coil sections, per unit length [H/m]
        self.setAttribute(self.Inputs, 'M_m', self.Auxiliary.overwrite_inductance_coil_sections)
        self.setAttribute(self.Inputs, 'HalfTurnToInductanceBlock', HalfTurnToInductanceBlock)
        self.setAttribute(self.Inputs, 'M_InductanceBlock_m', 0)  # This entry will tell LEDET to read the turn-to-turn inductances from the auxiliary .csv file


    def assignDefaultValuesWindings(self):
        '''
        Assign default values to LEDET variables defining coil windings parameters
        This is useful when defining a LEDET model of a conductor (case_model='conductor'), which does not need coil windings parameters
        '''

        self.setAttribute(self.Inputs, 'GroupToCoilSection', 1)
        self.setAttribute(self.Inputs, 'polarities_inGroup', 1)
        self.setAttribute(self.Inputs, 'nT', 1)
        self.setAttribute(self.Inputs, 'l_mag_inGroup', self.model_data.GeneralParameters.length_busbar)

        self.setAttribute(self.Inputs, 'alphasDEG', 0)
        self.setAttribute(self.Inputs, 'rotation_block', 0)
        self.setAttribute(self.Inputs, 'mirror_block', 0)
        self.setAttribute(self.Inputs, 'mirrorY_block', 0)

        self.setAttribute(self.Inputs, 'el_order_half_turns', 1)

        self.setAttribute(self.Inputs, 'iContactAlongWidth_From', 1)
        self.setAttribute(self.Inputs, 'iContactAlongWidth_To', 1)
        self.setAttribute(self.Inputs, 'iContactAlongHeight_From', 1)
        self.setAttribute(self.Inputs, 'iContactAlongHeight_To', 1)

        self.setAttribute(self.Inputs, 'M_m', 0.000001)
        self.setAttribute(self.Inputs, 'HalfTurnToInductanceBlock', 1)
        self.setAttribute(self.Inputs, 'M_InductanceBlock_m', 0.000001)
        self.setAttribute(self.Inputs, 'fL_I', np.array([0, 10000000000]))
        self.setAttribute(self.Inputs, 'fL_L', np.array([1, 1]))


    def localsParser(self, locals: dict):
        """
            Sets parameters in LEDET from 'locals' dictionary
            :param locals: dictionary with LEDET parameters
        """
        for attribute in locals:
            if attribute in self.Inputs.__annotations__:
                group = self.Inputs
            elif attribute in self.Options.__annotations__:
                group = self.Options
            elif attribute in self.Plots.__annotations__:
                group = self.Plots
            elif attribute in self.Variables.__annotations__:
                group = self.Variables
            else:
                continue

            tt = type(self.getAttribute(group, attribute))
            if tt == np.ndarray and isinstance(locals[attribute], list):
                self.setAttribute(group, attribute, np.array(locals[attribute]))
            elif tt == np.ndarray and not isinstance(locals[attribute], np.ndarray):
                self.setAttribute(group, attribute, np.array([locals[attribute]]))
            else:
                self.setAttribute(group, attribute, locals[attribute])


    def printVariableDescNameValue(self, variableGroup, variableLabels):
        """

           **Print variable description, variable name, and variable value**

           Function prints variable description, variable name, and variable value

           :param variableGroup: Dataclass containing all the attributes of the LEDET object
           [obsolete, but still supported: list of tuples; each tuple has two elements: the first element is a string defining
           the variable name, and the second element is either an integer, a float, a list, or a numpy.ndarray
           defining the variable value :type variableGroup: list :param variableLabels: dictionary assigning a
           description to each variable name]
           :type variableLabels: dataclass [obsolete, but still supported: dict]

           :return: None

           [Example for usage of obsolete dictionary-version]
            import numpy as np
            variableGroup = []
            variableGroup.append( ('x1', 12) )
            variableGroup.append( ('x2', 23.42) )
            variableGroup.append( ('x3', [2, 4, 6]) )
            variableGroup.append( ('x3', np.array([2, 4, 6])) )
            variableLabels = {'x1': '1st variable', 'x2': '2nd variable', 'x3': '3rd variable'}
            utils.printVariableDescNameValue(variableGroup, variableLabels)
            # >>> 					1st variable x1 12
            # >>> 					2nd variable x2 23.42
            # >>> 					3rd variable x3 [2, 4, 6]
            # >>> 					3rd variable x3 [2 4 6]

        """
        if(variableGroup == asdict(self.Inputs)):
            variableGroup = self.Inputs
        if (variableGroup == asdict(self.Options)):
            variableGroup = self.Options
        if (variableGroup == asdict(self.Plots)):
            variableGroup = self.Plots
        if (variableGroup == asdict(self.Variables)):
            variableGroup = self.Variables

        if(type(variableGroup) != dict):
            for k in variableGroup.__annotations__:
                if k == 'overwrite_f_internalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                if k == 'overwrite_f_externalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                print(variableLabels[k])
                print(k, self.getAttribute(variableGroup, k))
        else:
            for k in variableGroup:
                if k == 'overwrite_f_internalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                if k == 'overwrite_f_externalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
                print(variableLabels[k])
                print(k, variableGroup[k])


########################################################################################################################
# CLASS ENDS HERE - WHAT FOLLOWS IS A COPY/PASTE FROM PARAMETERS_LEDET TO INTEGRATE
########################################################################################################################


# class ParametersLEDET:
#     '''
#         Class of LEDET parameters
#     '''
#     def setAttribute(self, LEDETclass, attribute, value):
#         try:
#             setattr(LEDETclass, attribute, value)
#         except:
#             setattr(getattr(self, LEDETclass), attribute, value)
#
#     def getAttribute(self, LEDETclass, attribute):
#         try:
#             return getattr(LEDETclass, attribute)
#         except:
#             return getattr(getattr(self, LEDETclass), attribute)
#
#     def fillAttribute(self, LEDETclass, attribute, value):
#         imitate = self.getAttribute(LEDETclass, attribute)
#         if isinstance(imitate, np.ndarray) and isinstance(value, np.ndarray):
#             if imitate.shape != value.shape:
#                 imitate.resize(value.shape, refcheck=False)
#
#         idx_v = np.where(value != 0)
#         imitate[idx_v] = value[idx_v]
#         try:
#             setattr(LEDETclass, attribute, imitate)
#         except:
#             setattr(getattr(self, LEDETclass), attribute, imitate)
#
#
#     def __cpCu_nist_mat(self, T):
#         density = 8960
#         cpCu_perMass = np.zeros(T.size)
#         T[T < 4] = 4
#         idxT1 = np.where(T < 300)
#         idxT2 = np.where(T >= 300)
#         dc_a = -1.91844
#         dc_b = -0.15973
#         dc_c = 8.61013
#         dc_d = -18.996
#         dc_e = 21.9661
#         dc_f = -12.7328
#         dc_g = 3.54322
#         dc_h = -0.3797
#
#         logT1 = np.log10(T[idxT1])
#         tempVar = \
#         dc_a + dc_b * (logT1)**1 + dc_c * (logT1)**2 + dc_d * (logT1)**3 + \
#         dc_e * (logT1)**4 + dc_f * (logT1)**5 + dc_g * (logT1)** 6 + dc_h * (logT1)**7
#         cpCu_perMass[idxT1] = 10**tempVar
#
#         cpCu_perMass[idxT2]= 361.5 + 0.093 * T[idxT2]
#         cpCu = density * cpCu_perMass
#         return cpCu
#
#     def __rhoCu_nist(self, T, B, RRR, f_MR = 1):
#         B = abs(B)
#
#         idxLowB = np.where(B < 0.1)
#         idxHighB = np.where(B >= 0.1)
#
#         rho0 = 1.553e-8 / RRR
#         rhoi = 1.171e-17 * (T** 4.49) / (1 + 4.48e-7 * (T** 3.35) * np.exp(-(50. / T)** 6.428))
#         rhoiref = 0.4531 * rho0 * rhoi / (rho0 + rhoi)
#         rhcu = rho0 + rhoi + rhoiref
#         rhoCu = np.zeros(B.shape)
#         rhoCu[idxLowB] = rhcu[idxLowB]
#
#         lgs = 0.43429 * np.log(1.553E-8 * B[idxHighB] / rhcu[idxHighB])
#         polys = -2.662 + lgs * (0.3168 + lgs * (0.6229 + lgs  * (-0.1839 + lgs * 0.01827)))
#         corrs = (10.**polys)
#         rhoCu[idxHighB] = (1. + corrs * f_MR) * rhcu[idxHighB]
#         return rhoCu
#
#     def _rhoSS(self, T):
#         LimitValidityLow = 0
#         LimitValidityHigh = 300
#
#         fit_rho_SS_CERN = np.array([-6.16E-15, 3.52E-12, 1.72E-10, 5.43E-07]) / 1.0867
#         fit_rho_SS_CERN_linearExtrapolation = np.array([7.24E-10, 5.2887E-7]) / 1.0867
#
#         rhoSS = 0
#         if T < LimitValidityLow:
#             rhoSS = np.polyval(fit_rho_SS_CERN, LimitValidityLow)
#         elif T >= LimitValidityLow and T <= LimitValidityHigh:
#             rhoSS = np.polyval(fit_rho_SS_CERN, T)
#         elif T > LimitValidityHigh:
#             rhoSS = np.polyval(fit_rho_SS_CERN_linearExtrapolation, T)
#         return rhoSS
#
#     def __kCu_WiedemannFranz(self, rhoCu, T):
#         L0 = 2.45E-8
#         kCu = L0 * T / rhoCu
#         return kCu
#
#     def __kG10(self, T):
#         kG10 = np.zeros(T.size)
#         LimitValidity = 500
#         idxT1 = np.where(T <= LimitValidity)
#         idxT2 = np.where(T > LimitValidity)
#
#         a, b, c, d, e, f, g, h = -4.1236, 13.788, -26.068, 26.272, -14.663, 4.4954, -0.6905, 0.0397
#         logT = np.log10(T[idxT1])
#         logk = a + b * logT + c * logT**2 + d * logT**3 + e * logT**4 + f * logT**5 + g * logT**6 + h * logT**7
#         kG10[idxT1] = 10**logk
#
#         logLimitValidity = np.log10(LimitValidity)
#         logkLimitValidity = a + b * logLimitValidity + c * logLimitValidity**2 + d * logLimitValidity**3 + e * logLimitValidity**4 + \
#         f * logLimitValidity**5 + g * logLimitValidity**6 + h * logLimitValidity**7;
#         kG10[idxT2] = 10**logkLimitValidity
#         return kG10
#
#     def __cpG10(self, T):
#         density, a0, a1, a2, a3, a4, a5, a6, a7 = 1900, -2.4083, 7.6006, -8.2982, 7.3301, -4.2386, 1.4294, -0.24396, 0.015236
#         logT = np.log10(T)
#         p = 10**(a7 * ((logT)**7) + a6 * ((logT)**6) + a5 * ((logT)**5) + a4 * ((logT)**4) + a3 * ((logT)**3) + a2 * (
#                     (logT)**2) + a1 * ((logT)) + a0)
#         cpG10 = density * p
#         return cpG10
#
#     def __kKapton(self, T):
#         kKapton = np.zeros(T.size)
#         idxLow = np.where(T < 4.3)
#         if idxLow:
#             kKapton[idxLow[0]] = 0.010703 - 0.00161 * (4.3 - T[idxLow[0]])
#         idxHigh = np.where(T >= 4.3)
#         if idxHigh:
#             a, b, c, d, e, f, g, h = 5.73101, -39.5199, 79.9313, -83.8572, 50.9157, -17.9835, 3.42413, -0.27133
#             logT = np.log10(T)
#             logk = a + b * logT + c * logT**2 + d * logT**3 + e * logT**4 + f * logT**5 + g * logT**6 + h * logT**7
#             kKapton[idxHigh[0]] = 10**logk[idxHigh[0]]
#         return kKapton
#
#     def __cpKapton(self, T):
#         density, a0, a1, a2, a3, a4, a5, a6, a7 = 1420, -1.3684, 0.65892, 2.8719, 0.42651, -3.0088, 1.9558, -0.51998, 0.051574
#         logT = np.log10(T)
#         p = 10**(a7 * ((logT)**7) + a6 * ((logT)**6) + a5 * ((logT)**5) + a4 * ((logT)**4) + a3 * (
#                     (logT)**3) + a2 * ((logT)**2) + a1 * ((logT)) + a0)
#         cpKapton = density * p
#         return cpKapton
#
#     def __cpNbTi_cudi_mat(self, T, B):
#         Tc0 = 9.2
#         Bc20 = 14.5
#         alpha = .59
#         B[B>= Bc20] = Bc20-10E-4
#
#         Tc = Tc0 * (1 - B / Bc20)**alpha
#         cpNbTi = np.zeros(T.size)
#
#         idxT1 = np.where(T <= Tc)
#         idxT2 = np.where((T > Tc) & (T <= 20.0))
#         idxT3 = np.where((T > 20) & (T <= 50))
#         idxT4 = np.where((T > 50) & (T <= 175))
#         idxT5 = np.where((T > 175) & (T <= 500))
#         idxT6 = np.where((T > 500) & (T <= 1000))
#         idxT7 = np.where(T > 1000)
#
#         p1 = [0.00000E+00,    4.91000E+01,   0.00000E+00,   6.40000E+01,  0.00000E+00]
#         p2 = [0.00000E+00,   1.62400E+01,   0.00000E+00,  9.28000E+02,   0.00000E+00]
#         p3 = [-2.17700E-01,   1.19838E+01,   5.53710E+02, - 7.84610E+03,  4.13830E+04]
#         p4 = [-4.82000E-03,  2.97600E+00, -7.16300E+02,  8.30220E+04,  -1.53000E+06]
#         p5 = [-6.29000E-05, 9.29600E-02, -5.16600E+01,  1.37060E+04,  1.24000E+06]
#         p6 = [0.00000E+00, 0.00000E+00,  -2.57000E-01,  9.55500E+02,  2.45000E+06]
#         p7 = [0, 0, 0, 0, 3.14850E+06]
#
#         cpNbTi[idxT1] = p1[0] * T[idxT1]**4 + p1[1] * T[idxT1]**3 + p1[2] * T[idxT1]**2 + p1[3] * T[idxT1] + p1[4]
#         cpNbTi[idxT2] = p2[0] * T[idxT2]**4 + p2[1] * T[idxT2]**3 + p2[2] * T[idxT2]**2 + p2[3] * T[idxT2] + p2[4]
#         cpNbTi[idxT3] = p3[0] * T[idxT3]**4 + p3[1] * T[idxT3]**3 + p3[2] * T[idxT3]**2 + p3[3] * T[idxT3] + p3[4]
#         cpNbTi[idxT4] = p4[0] * T[idxT4]**4 + p4[1] * T[idxT4]**3 + p4[2] * T[idxT4]**2 + p4[3] * T[idxT4] + p4[4]
#         cpNbTi[idxT5] = p5[0] * T[idxT5]**4 + p5[1] * T[idxT5]**3 + p5[2] * T[idxT5]**2 + p5[3] * T[idxT5] + p5[4]
#         cpNbTi[idxT6] = p6[0] * T[idxT6]**4 + p6[1] * T[idxT6]**3 + p6[2] * T[idxT6]**2 + p6[3] * T[idxT6] + p6[4]
#         cpNbTi[idxT7] = p7[0] * T[idxT7]**4 + p7[1] * T[idxT7]**3 + p7[2] * T[idxT7]**2 + p7[3] * T[idxT7] + p7[4]
#         return cpNbTi
#
#     def __cpNb3Sn_alternative_mat(self, T, B, Tc0_Nb3Sn, Bc20_Nb3Sn):
#         B[B < .001] = 0.001
#         cpNb3Sn = np.zeros(T.shape)
#         alpha = .59
#         Tc = Tc0_Nb3Sn * (1 - B / Bc20_Nb3Sn)** alpha
#         density = 8950.0 # [kg / m ^ 3]
#
#         idxT0 = np.where(T <= Tc)
#         idxT1 = np.where((T > Tc) & (T <= 20))
#         idxT2 = np.where((T > 20) & (T <= 400))
#         idxT3 = np.where(T > 400)
#
#
#         betaT = 1.241E-3 # [J / K ^ 4 / kg]
#         gammaT = .138 # [J / K ^ 2 / kg]
#
#         if len(B) > 1 and len(Tc0_Nb3Sn) > 1:
#             cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn[idxT0]** 2) * T[idxT0]** 3 + gammaT* B[idxT0] / Bc20_Nb3Sn[idxT0] * T[idxT0]
#         elif len(B) > 1:
#             cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn** 2) * T[idxT0]**3 + gammaT * B[idxT0] / Bc20_Nb3Sn * T[idxT0]
#         elif len(Tc0_Nb3Sn) > 1:
#             cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn[idxT0]** 2) * T[idxT0]**3 + gammaT * B / Bc20_Nb3Sn[idxT0] * T[idxT0]
#         elif len(B) == 1 and len(Tc0_Nb3Sn) == 1:
#             cpNb3Sn[idxT0] = (betaT + 3 * gammaT / Tc0_Nb3Sn**2) * T[idxT0]**3 + gammaT * B / Bc20_Nb3Sn * T[idxT0]
#
#         cpNb3Sn[idxT1] = betaT * T[idxT1]**3 + gammaT * T[idxT1]
#         polyFit_20K_400K = [0.1662252, -0.6827738, -6.3977, 57.48133, -186.90995, 305.01434, -247.44839, 79.78547]
#         logT = np.log10(T[idxT2])
#         logCp2 = np.polyval(polyFit_20K_400K, logT)
#         cpNb3Sn[idxT2] = 10** logCp2
#
#         log400K = np.log10(400)
#         logCp400K = np.polyval(polyFit_20K_400K, log400K)
#         cpNb3Sn[idxT3] = 10**logCp400K
#         cpNb3Sn = cpNb3Sn * density
#         return cpNb3Sn
#
#     def __Jc_Nb3Sn_Summer(self, T, B, Jc_Nb3Sn0, Tc0_Nb3Sn, Bc20_Nb3Sn):
#         if type(T)== int or type(T)== float:
#             T = np.repeat(T, len(Jc_Nb3Sn0)).astype(float)
#
#         B[abs(B) < .001] = 0.001
#         T[T < 0.001] = 0.001
#         f_T_T0 = T / Tc0_Nb3Sn
#         f_T_T0[f_T_T0 > 1] = 1
#         Bc2 = Bc20_Nb3Sn * (1 - f_T_T0**2) * (1 - 0.31 * f_T_T0**2 * (1 - 1.77 * np.log(f_T_T0)))
#         f_B_Bc2 = B / Bc2
#         f_B_Bc2[f_B_Bc2 > 1] = 1
#         Jc_T_B = Jc_Nb3Sn0 / np.sqrt(B) * (1 - f_B_Bc2)**2 * (1 - f_T_T0** 2)**2
#         return Jc_T_B
#
#     def __Tc_Tcs_Nb3Sn_approx(self, J, B, Jc_Nb3Sn0, Tc0_Nb3Sn, Bc20_Nb3Sn):
#         J = abs(J)
#         B = abs(B)
#
#         f_B_Bc2 = B / Bc20_Nb3Sn
#         f_B_Bc2[f_B_Bc2 > 1] = 1
#         Tc = Tc0_Nb3Sn * (1 - f_B_Bc2)**.59
#
#         Jc0 = self.__Jc_Nb3Sn_Summer(0, B, Jc_Nb3Sn0, Tc0_Nb3Sn, Bc20_Nb3Sn)
#         f_J_Jc0 = J/ Jc0
#         f_J_Jc0[f_J_Jc0 > 1] = 1
#
#         Tcs = (1 - f_J_Jc0) * Tc
#         return [Tc, Tcs]
#
#     def _obtainThermalConnections(self):
#         # Calculate group to which each half-turn belongs
#         nHalfTurnsDefined = len(self.Inputs.HalfTurnToInductanceBlock)
#         indexTstart = np.hstack([1, 1 + np.cumsum(self.Inputs.nT[:-1])]).astype(int)
#         indexTstop = np.cumsum(self.Inputs.nT).astype(int)
#         HalfTurnToGroup = np.zeros((1, nHalfTurnsDefined), dtype=int)
#         HalfTurnToGroup = HalfTurnToGroup[0]
#         for g in range(1, len(self.Inputs.nT) + 1):
#             HalfTurnToGroup[indexTstart[g - 1] - 1:indexTstop[g - 1]] = g
#
#         # Obtain all thermal connections of each turn and store them in dictionaries for width and height
#         # First width
#         th_con_w = {}
#         for i in range(1, len(self.Inputs.HalfTurnToInductanceBlock) + 1):
#             con_list = []
#             iWidthFrom = np.where(self.Inputs.iContactAlongWidth_From == i)
#             if iWidthFrom: con_list = con_list + self.Inputs.iContactAlongWidth_To[iWidthFrom[0]].astype(int).tolist()
#             iWidthTo = np.where(self.Inputs.iContactAlongWidth_To == i)
#             if iWidthTo: con_list = con_list + self.Inputs.iContactAlongWidth_From[iWidthTo[0]].astype(int).tolist()
#             th_con_w[str(i)] = con_list
#
#         # Then height
#         th_con_h = {}
#         for i in range(1, len(self.Inputs.HalfTurnToInductanceBlock) + 1):
#             con_list = []
#             iHeightFrom = np.where(self.Inputs.iContactAlongHeight_From == i)
#             if iHeightFrom: con_list = con_list + self.Inputs.iContactAlongHeight_To[iHeightFrom[0]].astype(
#                 int).tolist()
#             iHeightTo = np.where(self.Inputs.iContactAlongHeight_To == i)
#             if iHeightTo: con_list = con_list + self.Inputs.iContactAlongHeight_From[iHeightTo[0]].astype(int).tolist()
#             th_con_h[str(i)] = con_list
#         return [HalfTurnToGroup, th_con_w, th_con_h]
#
#     def __calculateTransversalDelay(self, cp, kIns, Tc, Tcs, T_bath):
#         [HalfTurnToGroup, th_con_w, th_con_h] = self._obtainThermalConnections()
#         # Use dictionaries to calculate the transversal quench delay into each direction based on respective properties
#         delta_t_w = {}
#         delta_t_h = {}
#         for i in range(1,len(self.Inputs.HalfTurnToInductanceBlock)+1):
#             con = th_con_h[str(i)]
#             delta_t_h_temp = []
#             for k in range(len(con)):
#                 idx_con1 = HalfTurnToGroup[k-1]-1
#                 idx_con2 = HalfTurnToGroup[con[k]-1]-1
#                 T_temp = 1
#                 delta_t = cp[idx_con2] / kIns[idx_con2] * (
#                             self.Inputs.wBare_inGroup[idx_con2] + 2 * self.Inputs.wIns_inGroup[idx_con2]) \
#                           * (self.Inputs.wIns_inGroup[idx_con2] + self.Inputs.wIns_inGroup[idx_con1]) * T_temp
#                 delta_t_h_temp.append(delta_t)
#             delta_t_h[str(i)] = delta_t_h_temp
#
#             con = th_con_w[str(i)]
#             delta_t_w_temp = []
#             for k in range(len(con)):
#                 idx_con1 = HalfTurnToGroup[k - 1] - 1
#                 idx_con2 = HalfTurnToGroup[con[k]-1]-1
#                 T_temp = (Tcs[idx_con2]-T_bath)/(Tc[idx_con1]-(Tcs[idx_con2]+T_bath)/2)
#                 delta_t = cp[idx_con2] / kIns[idx_con2] * (
#                             self.Inputs.hBare_inGroup[idx_con2] + 2 * self.Inputs.hIns_inGroup[idx_con2]) \
#                           * (self.Inputs.hIns_inGroup[idx_con2] + self.Inputs.hIns_inGroup[idx_con1]) * T_temp
#                 delta_t_w_temp.append(delta_t)
#             delta_t_w[str(i)] = delta_t_w_temp
#
#         return [HalfTurnToGroup, th_con_h, delta_t_h, th_con_w, delta_t_w]
#
#     def _quenchPropagationVelocity(self, I, B, T_bath, cable):
#         # Calculate Quench propagation velocity
#         L0 = 2.44E-08
#         A_CableBare = cable.A_CableInsulated * (cable.f_SC + cable.f_ST)
#         f_SC_inStrand = cable.f_SC / (cable.f_SC + cable.f_ST)
#         f_ST_inStrand = cable.f_ST / (cable.f_SC + cable.f_ST)
#         I = abs(I)
#         J_op = I / A_CableBare
#
#         idxNbTi = np.where(np.repeat(self.Inputs.SCtype_inGroup,self.Inputs.nT.astype(int)) == 1)[0]
#         idxNb3Sn = np.where(np.repeat(self.Inputs.SCtype_inGroup,self.Inputs.nT.astype(int)) == 2)[0]
#         idxCu_ST = np.where(np.repeat(self.Inputs.STtype_inGroup,self.Inputs.nT.astype(int)) == 1)[0]
#
#         Tc = np.zeros(B.shape)
#         Tcs = np.zeros(B.shape)
#         if len(idxNbTi)>0:
#             Tc[idxNbTi] = cable.Tc0_NbTi[idxNbTi] * (1 - B / cable.Bc20_NbTi[idxNbTi]) ** cable.alpha_NbTi
#             Tcs[idxNbTi] = (1 - I / (cable.c1_Ic_NbTi[idxNbTi] + cable.c2_Ic_NbTi[idxNbTi] * B[idxNbTi])) * Tc[idxNbTi]
#         if len(idxNb3Sn) > 0:
#             [Tc[idxNb3Sn], Tcs[idxNb3Sn]] = self.__Tc_Tcs_Nb3Sn_approx(I / cable.A_SC[idxNb3Sn], B[idxNb3Sn],
#                                                                 cable.Jc_Nb3Sn0[idxNb3Sn], cable.Tc0_Nb3Sn[idxNb3Sn],
#                                                                 cable.Bc20_Nb3Sn[idxNb3Sn])
#
#         Ts = (Tcs + Tc) / 2
#         cp_ST = np.zeros(B.shape)
#         cp_ST[idxCu_ST] = self.__cpCu_nist_mat(Ts[idxCu_ST])
#         cp_SC = np.zeros(B.shape)
#         if len(idxNbTi) > 0:
#             cp_SC[idxNbTi] = self.__cpNbTi_cudi_mat(Ts[idxNbTi], B[idxNbTi])
#         if len(idxNb3Sn) > 0:
#             cp_SC[idxNb3Sn] = self.__cpNb3Sn_alternative_mat(Ts[idxNb3Sn], B[idxNb3Sn], cable.Tc0_Nb3Sn[idxNb3Sn], cable.Bc20_Nb3Sn[idxNb3Sn])
#         cp = cp_ST * f_ST_inStrand + cp_SC * f_SC_inStrand
#         vQ = J_op / cp * ((L0 * Ts) / (Ts - T_bath))**0.5
#         idxInfQuenchVelocity=np.where(Tcs <= T_bath)
#         vQ[idxInfQuenchVelocity]=1E6
#
#         ### Calculate MPZ
#         rhoCu = np.zeros(A_CableBare.shape)
#         kCu = np.zeros(A_CableBare.shape)
#         RRR = np.repeat(self.Inputs.RRR_Cu_inGroup, self.Inputs.nT.astype(int))
#         rhoCu[idxCu_ST] = self.__rhoCu_nist(Ts[idxCu_ST], B[idxCu_ST], RRR[idxCu_ST])
#         kCu[idxCu_ST] = self.__kCu_WiedemannFranz(rhoCu[idxCu_ST], Ts[idxCu_ST])
#         l = np.zeros(A_CableBare.shape)
#         l[idxCu_ST] = np.sqrt((2 * kCu[idxCu_ST] * (Tc[idxCu_ST] - T_bath)) / (J_op[idxCu_ST]** 2 * rhoCu[idxCu_ST]))
#
#         # Calculate thermal conductivity of insulations
#         idxKapton = np.where(self.Inputs.insulationType_inGroup == 2, 1, 0)
#         idxKapton = np.where(np.repeat(idxKapton, self.Inputs.nT.astype(int))==1)[0]
#         idxG10 = np.where(self.Inputs.insulationType_inGroup == 1, 1, 0)
#         idxG10 = np.where(np.repeat(idxG10, self.Inputs.nT.astype(int))==1)[0]
#         kIns = np.zeros(Ts.size)
#         kIns[idxKapton] = self.__kKapton(Ts[idxKapton])
#         kIns[idxG10] = self.__kG10(Ts[idxG10])
#         cpIns = np.zeros(Ts.size)
#         cpIns[idxKapton] = self.__cpKapton(Ts[idxKapton])
#         cpIns[idxG10] = self.__cpG10(Ts[idxG10])
#         cp_full = (cp* (A_CableBare/cable.A_CableInsulated) + cpIns*(1-A_CableBare/cable.A_CableInsulated))/2
#
#         ### Calculate delta T transversal
#         [HalfTurnToGroup, th_con_h, delta_t_h, th_con_w, delta_t_w] = self.__calculateTransversalDelay(cp_full, kIns, Tc, Tcs, T_bath)
#         return [vQ, l, HalfTurnToGroup, th_con_h, delta_t_h, th_con_w, delta_t_w]
#
#     def __reorderROXIEFiles(self, ROXIE_File):
#         orderedROXIE = []
#         for i in range(len(ROXIE_File)-1):
#             prefix = 'E'+str(i)
#             for j in range(len(ROXIE_File)):
#                 if prefix in ROXIE_File[j]:
#                     orderedROXIE.append(ROXIE_File[j])
#         for j in range(len(ROXIE_File)):
#             if 'All_WithIron_WithSelfField' in ROXIE_File[j]:
#                 orderedROXIE.append(ROXIE_File[j])
#         return orderedROXIE
#
#     def _acquireBField(self, ROXIE_File):
#         if ROXIE_File.endswith('.map2d'):
#             ROXIE_File = [ROXIE_File]
#             N = 1
#         else:
#             ROXIE_File1 = [f for f in os.listdir(ROXIE_File) if os.path.isfile(os.path.join(ROXIE_File, f))]
#             ROXIE_File1 = [f for f in ROXIE_File1 if f.endswith('.map2d')]
#             ROXIE_File1 = [f for f in ROXIE_File1 if 'WithIron' in f]
#             ROXIE_File1 = [f for f in ROXIE_File1 if not 'ROXIE' in f]
#             for i in range(len(ROXIE_File1)):
#                 ROXIE_File1[i] = os.path.join(ROXIE_File, ROXIE_File1[i])
#             ROXIE_File = ROXIE_File1
#             ROXIE_File = self.__reorderROXIEFiles(ROXIE_File)
#             N = len(ROXIE_File)
#             if N>1:
#                 print('Reading ', N, ' Field maps. This may take a while.')
#             else:
#                 print('Reading Field map.')
#
#         for i in trange(N, file=sys.stdout, desc='Field maps'):
#             Inom = self.Options.Iref
#             reader = csv.reader(open(ROXIE_File[i]))
#             B_Field = np.array([])
#             stack = 0
#             for row in reader:
#                 if not row: continue
#                 row_s = np.array(row[0].split())
#                 if not stack:
#                     B_Field = np.array(row_s[1:])
#                     stack = 1
#                 else:
#                     B_Field = np.vstack((B_Field, np.array(row_s)))
#             B_Field = B_Field[1:].astype(float)
#             if i == 0:
#                 BX = (B_Field[:, 5].transpose()/ Inom)
#                 BY = (B_Field[:, 6].transpose()/ Inom)
#             elif i == N-1:
#                 BX_All = B_Field[:, 5].transpose()
#                 BY_All = B_Field[:, 6].transpose()
#             else:
#                 BX = BX + (B_Field[:, 5].transpose() / Inom)
#                 BY = BY + (B_Field[:, 6].transpose() / Inom)
#         f_mag = (BX** 2 + BY** 2) ** 0.5
#         if N>1:
#             B_E_All = (BX_All** 2 + BY_All** 2) ** 0.5
#             peakB_Superposition = max(f_mag * Inom)
#             peakB_Real = max(B_E_All)
#             f_peakReal_peakSuperposition = peakB_Real / peakB_Superposition
#         else: f_peakReal_peakSuperposition = 1
#
#         B = f_mag*self.Inputs.I00*f_peakReal_peakSuperposition
#
#         B[B > 10E6]=10E-6
#         return B
#
#     def __repeatCable(self, cable):
#         nT = self.Inputs.nT
#         nT = nT.astype(int)
#         newCable = Cable()
#         for attribute in cable.__annotations__:
#             if attribute == 'alpha_NbTi': continue
#             x = np.ndarray([])
#             x = getattr(cable, attribute)
#             x = np.repeat(x, nT)
#             setattr(newCable, attribute, x)
#         return newCable
#
#     def calculateQuenchDetectionTime(self, Type, B, vQ_iStartQuench, lengthHotSpot_iStartQuench, HalfTurnToGroup, th_con_h, delta_t_h, th_con_w, delta_t_w, uQuenchDetectionThreshold = 0.1):
#         if not (Type=='Short' or Type=='Long'):
#             print("Don't understand type of quench detection time calculation. Please choose either 'Short' or 'Long'[incl. transversal propagation]")
#             print("Let's continue with 'Short'.")
#             Type = 'Short'
#         # Calculate resistance of each turn at T=10 K
#         rho_Cu_10K = 1.7E-10  # [Ohm*m] Approximate Cu resistivity at T=10 K, B=0, for RRR=100
#         rho_Cu_10K_B = 4E-11  # [Ohm*m/T] Approximate Cu magneto-resistivity factor
#         Iref = self.Options.Iref
#         nStrands_inGroup = self.Inputs.nStrands_inGroup
#         ds_inGroup = self.Inputs.ds_inGroup
#         f_SC_strand_inGroup = self.Inputs.f_SC_strand_inGroup
#         nHalfTurns = len(vQ_iStartQuench)
#
#         tQuenchDetection = []
#         r_el_m = np.zeros((nHalfTurns,))
#         for ht in range(1, nHalfTurns + 1):
#             current_group = HalfTurnToGroup[ht - 1]
#             mean_B = B / Iref * self.Inputs.I00  # average magnetic field in the current half-turn
#             rho_mean = rho_Cu_10K + rho_Cu_10K_B * mean_B[ht-1]  # average resistivity in the current half-turn
#             cross_section = nStrands_inGroup[current_group - 1] * np.pi / 4 * ds_inGroup[current_group - 1] ** 2 * (1 - f_SC_strand_inGroup[current_group - 1])
#             r_el_m[ht - 1] = rho_mean / cross_section # Electrical resistance per unit length
#             if Type == 'Short':
#                 UQD_i = (self.Inputs.I00 * r_el_m[ht - 1] * lengthHotSpot_iStartQuench[ht - 1])
#                 tQD = (uQuenchDetectionThreshold - UQD_i) / (vQ_iStartQuench[ht - 1] * r_el_m[ht - 1] * self.Inputs.I00)
#                 tQuenchDetection = np.hstack([tQuenchDetection, np.array(tQD)])
#
#         r_el_m = r_el_m.transpose()
#         if Type == 'Long':
#             for ht in range(1, nHalfTurns + 1):
#                 for ht in range(1, nHalfTurns + 1):
#                     # Approximate time to reach the quench detection threshold
#                     UQD_i = (self.Inputs.I00 * r_el_m[ht - 1] * lengthHotSpot_iStartQuench[ht - 1])
#                     tQD = (uQuenchDetectionThreshold - UQD_i) / (vQ_iStartQuench[ht - 1] * r_el_m[ht - 1] * self.Inputs.I00)
#                     delay = np.concatenate((np.array(delta_t_w[str(ht)]), np.array(delta_t_h[str(ht)])), axis=None)
#                     th_con = np.concatenate((np.array(th_con_w[str(ht)]), np.array(th_con_h[str(ht)])), axis=None).astype(int)
#                     tQD_i = tQD
#                     t_i0 = 0
#                     t_i1 = 0
#                     idx_turns = np.array([ht - 1])
#                     quenched_turns = [ht]
#                     delay[delay > tQD_i] = 9999
#
#                     while np.any(delay < 999):
#                         idx = np.argmin(delay)
#                         if th_con[idx] in quenched_turns:
#                             delay[idx] = 9999
#                             continue
#                         else:
#                             quenched_turns.append(int(th_con[idx]))
#                         UQD_i = UQD_i + np.sum(self.Inputs.I00 * r_el_m[idx_turns] * (t_i1 - t_i0) * vQ_iStartQuench[idx_turns])
#                         idx_turns = np.append(idx_turns, int(th_con[idx] - 1))
#                         t_i1 = delay[idx]
#                         tQD_i = (uQuenchDetectionThreshold - UQD_i) / (
#                                 np.sum(vQ_iStartQuench[idx_turns] * r_el_m[idx_turns] * self.Inputs.I00))
#                         t_i0 = t_i1
#                         delay = np.concatenate((delay, np.array(delta_t_w[str(int(th_con[idx]))] + t_i1),
#                                                     np.array(delta_t_h[str(int(th_con[idx]))] + t_i1)), axis=None)
#                         th_con = np.concatenate((th_con, np.array(th_con_w[str(int(th_con[idx]))]),
#                                                  np.array(th_con_h[str(int(th_con[idx]))])),axis=None)
#                         delay[delay > tQD_i] = 9999
#                         delay[idx] = 9999
#                 tQuenchDetection = np.hstack([tQuenchDetection, np.array(tQD_i)])
#         print('Minimum quench detection time would be {} ms [{} calculation]'.format(round(min(tQuenchDetection),3)*1000, Type))
#         return min(tQuenchDetection)
#
#     def getBField(self, ROXIE_File):
#         B = self._acquireBField(ROXIE_File)
#         strandCount = 0
#         GroupCount = 0
#         nStrands_inGroup = self.Inputs.nStrands_inGroup
#         ds_inGroup = self.Inputs.ds_inGroup
#         if any(nStrands_inGroup % 2 == 1) and any(nStrands_inGroup != 1):
#             for g in range(len(self.Inputs.nT)):
#                 if (nStrands_inGroup[g] % 2 == 1) & (nStrands_inGroup[g] > 1):
#                     ds_inGroup[g] = ds_inGroup[g] * np.sqrt(nStrands_inGroup[g] / (nStrands_inGroup[g] - 1))
#                     nStrands_inGroup[g] = nStrands_inGroup[g] - 1
#
#         Bcopy = np.zeros(int(sum(self.Inputs.nT)))
#         for i in range(int(sum(self.Inputs.nT))):
#             Bcopy[i] = sum(B[int(strandCount):int(strandCount + nStrands_inGroup[GroupCount])]) / nStrands_inGroup[
#                 int(GroupCount)]
#             TurnSum = sum(self.Inputs.nT[0:GroupCount + 1])
#             strandCount = strandCount + nStrands_inGroup[GroupCount]
#             if i > TurnSum: GroupCount = GroupCount + 1
#         return Bcopy
#
#     def adjust_vQ(self, ROXIE_File, Transversaldelay  = False, ManualB = '', CurrentsInCoilsections = []):
#         cable = Cable()
#         cable.A_CableInsulated = (self.Inputs.wBare_inGroup+2*self.Inputs.wIns_inGroup) \
#                                * (self.Inputs.hBare_inGroup+2*self.Inputs.hIns_inGroup)
#         if len(ManualB)==0: B = self._acquireBField(ROXIE_File)
#         else: B = ManualB
#
#         if max(self.Inputs.nStrands_inGroup) > 1:
#             strandCount = 0
#             GroupCount = 0
#             nStrands_inGroup = self.Inputs.nStrands_inGroup
#             ds_inGroup = self.Inputs.ds_inGroup
#             if any(nStrands_inGroup % 2 == 1) and any(nStrands_inGroup != 1):
#                 for g in range(len(self.Inputs.nT)):
#                     if (nStrands_inGroup[g] % 2 == 1) & (nStrands_inGroup[g] > 1):
#                         ds_inGroup[g] = ds_inGroup[g] * np.sqrt(nStrands_inGroup[g] / (nStrands_inGroup[g] - 1))
#                         nStrands_inGroup[g] = nStrands_inGroup[g] - 1
#             if len(ManualB)==0:
#                 Bcopy = np.zeros(int(sum(self.Inputs.nT)))
#                 for i in range(int(sum(self.Inputs.nT))):
#                     Bcopy[i] = sum(B[int(strandCount):int(strandCount+nStrands_inGroup[GroupCount])])/nStrands_inGroup[int(GroupCount)]
#                     TurnSum = sum(self.Inputs.nT[0:GroupCount+1])
#                     strandCount = strandCount + nStrands_inGroup[GroupCount]
#                     if i>TurnSum: GroupCount = GroupCount + 1
#                 B = Bcopy
#
#             cable.f_SC = self.Inputs.f_SC_strand_inGroup * \
#                          (nStrands_inGroup* (np.pi/4)*(ds_inGroup**2)) / cable.A_CableInsulated
#             cable.f_ST = (1 - self.Inputs.f_SC_strand_inGroup) * \
#                          (nStrands_inGroup* (np.pi/4)*(ds_inGroup**2)) / cable.A_CableInsulated
#         else:
#             cable.f_SC = self.Inputs.f_SC_strand_inGroup * \
#                          (self.Inputs.wBare_inGroup * self.Inputs.hBare_inGroup) / cable.A_CableInsulated
#             cable.f_ST = (1 - self.Inputs.f_SC_strand_inGroup) * \
#                          (self.Inputs.wBare_inGroup * self.Inputs.hBare_inGroup) / cable.A_CableInsulated
#
#         T_bath = self.Inputs.T00
#         cable.A_SC =cable.A_CableInsulated * cable.f_SC
#         cable.SCtype = self.Inputs.SCtype_inGroup
#         cable.STtype = self.Inputs.STtype_inGroup
#         cable.Tc0_NbTi = self.Inputs.Tc0_NbTi_ht_inGroup
#         cable.Bc20_NbTi = self.Inputs.Bc2_NbTi_ht_inGroup
#         cable.c1_Ic_NbTi = self.Inputs.c1_Ic_NbTi_inGroup
#         cable.c2_Ic_NbTi = self.Inputs.c2_Ic_NbTi_inGroup
#         cable.alpha_NbTi = .59
#         cable.Jc_Nb3Sn0 = self.Inputs.Jc_Nb3Sn0_inGroup
#         cable.Tc0_Nb3Sn = self.Inputs.Tc0_Nb3Sn_inGroup
#         cable.Bc20_Nb3Sn = self.Inputs.Bc2_Nb3Sn_inGroup
#         cable = self.__repeatCable(cable)
#
#         th_con_h = []
#         th_con_w = []
#         if len(CurrentsInCoilsections)>0:
#             if np.max(self.Inputs.GroupToCoilSection) != len(CurrentsInCoilsections):
#                 print('You assigned ', len(CurrentsInCoilsections),' currents in the coilsections, but there are ',
#                       np.max(self.Inputs.GroupToCoilSection), ' Coil-sections. Abort!')
#                 return
#
#             vQ_copy = np.linspace(0, len(cable.A_CableInsulated), len(cable.A_CableInsulated))
#             TurnToCoilSection = np.repeat(self.Inputs.GroupToCoilSection, self.Inputs.nT.astype(int))
#             for i in range(len(CurrentsInCoilsections)):
#                 I = CurrentsInCoilsections[i]
#                 B_copy = B/ self.Inputs.I00 * I
#                 [vQ, l, HalfTurnToGroup, th_con_h, delta_t_h, th_con_w, delta_t_w] = \
#                     self._quenchPropagationVelocity(I, B_copy, T_bath, cable)
#
#                 idx_cs = np.where(TurnToCoilSection == i+1)[0]
#                 vQ_copy[idx_cs] = vQ[idx_cs]
#             vQ = vQ_copy
#         else:
#             I = self.Inputs.I00
#             [vQ, l, HalfTurnToGroup, th_con_h, delta_t_h, th_con_w, delta_t_w] = self._quenchPropagationVelocity(I, B, T_bath, cable)
#
#         # self.setAttribute(getattr(self, "Inputs"), "vQ_iStartQuench", vQ)
#         # self.setAttribute(getattr(self, "Inputs"), "lengthHotSpot_iStartQuench", l)
#         self.setAttribute(getattr(self, "Inputs"), "fScaling_vQ_iStartQuench", np.ones(len(vQ)))
#
#         if Transversaldelay:
#             if len(CurrentsInCoilsections)>0:
#                 print('Multiple currents in the coilsections are not supported for calculation of quench detection times. I use I_nom.')
#             tQD = self.calculateQuenchDetectionTime(Transversaldelay, B, vQ, l, HalfTurnToGroup, th_con_h, delta_t_h,
#                                               th_con_w, delta_t_w, uQuenchDetectionThreshold = 0.1)
#             return [l, tQD]
#         else:
#             return vQ, l
#
#     def adjust_vQ_QuenchHeater(self, th_con_h, th_con_w, NHeatingStations):
#         idx_turns_2x = np.array([])
#         activatedStrips = np.where(self.Inputs.tQH < 999)[0]+1
#         idx_turns_2x = np.append(idx_turns_2x, activatedStrips)
#
#         for i in activatedStrips:
#             idx_new_turns = np.where(self.Inputs.iQH_toHalfTurn_From == i)[0]
#             idx_turns_2x = np.append(idx_turns_2x, self.Inputs.iQH_toHalfTurn_To[idx_new_turns])
#         idx_turns_2x = idx_turns_2x.astype(int)
#         for j in range(5):
#             for i in idx_turns_2x:
#                 if str(i) in th_con_h.keys():
#                     for k in th_con_h[str(i)]:
#                         if k not in idx_turns_2x:
#                             idx_turns_2x = np.append(idx_turns_2x, k)
#                 if str(i) in th_con_w.keys():
#                     for k in th_con_w[str(i)]:
#                         if k not in idx_turns_2x:
#                             idx_turns_2x = np.append(idx_turns_2x, k)
#         idx_turns_2x = idx_turns_2x.astype(int)-1
#         self.Inputs.vQ_iStartQuench[idx_turns_2x] = self.Inputs.vQ_iStartQuench[idx_turns_2x] * NHeatingStations * 2
#         if len(self.Inputs.lengthHotSpot_iStartQuench) != len(self.Inputs.vQ_iStartQuench):
#             self.Inputs.lengthHotSpot_iStartQuench = np.array([0.01]*len(self.Inputs.vQ_iStartQuench))
#         if type(self.Inputs.lengthHotSpot_iStartQuench) != np.ndarray:
#             self.Inputs.lengthHotSpot_iStartQuench = np.array(self.Inputs.lengthHotSpot_iStartQuench)
#         #self.Inputs.lengthHotSpot_iStartQuench[idx_turns_2x] = np.array([0.01*NHeatingStations]*len(idx_turns_2x))
#         self.Inputs.lengthHotSpot_iStartQuench[idx_turns_2x] = np.array([self.Inputs.l_magnet*self.Inputs.f_QH[0]]*len(idx_turns_2x))
#         return
#
#     def setConductorResistanceFraction(self, f_RRR1_Cu_inGroup=0.3, f_RRR2_Cu_inGroup=0.3, f_RRR3_Cu_inGroup=0.2,
#               RRR1_Cu_inGroup=100, RRR2_Cu_inGroup=100, RRR3_Cu_inGroup=100):
#         """
#             **enables and sets values for resistance parameters. Afterwards, will be included when writing to file**
#         """
#
#         print('Conductor resistance fraction enabled.')
#
#         self.enableConductorResistanceFraction = True
#
#         self.Inputs.f_RRR1_Cu_inGroup = f_RRR1_Cu_inGroup
#         self.Inputs.f_RRR2_Cu_inGroup = f_RRR2_Cu_inGroup
#         self.Inputs.f_RRR3_Cu_inGroup = f_RRR3_Cu_inGroup
#
#         self.Inputs.RRR1_Cu_inGroup = RRR1_Cu_inGroup
#         self.Inputs.RRR2_Cu_inGroup = RRR2_Cu_inGroup
#         self.Inputs.RRR3_Cu_inGroup = RRR3_Cu_inGroup
#
#         self.descriptionsInputs['f_RRR1_Cu_inGroup'] = 'Description ...'
#         self.descriptionsInputs['f_RRR2_Cu_inGroup'] = 'Description ...'
#         self.descriptionsInputs['f_RRR3_Cu_inGroup'] = 'Description ...'
#
#         self.descriptionsInputs['RRR1_Cu_inGroup'] = 'Description ...'
#         self.descriptionsInputs['RRR2_Cu_inGroup'] = 'Description ...'
#         self.descriptionsInputs['RRR3_Cu_inGroup'] = 'Description ...'
#
#     def setHeliumFraction(self, PercentVoids):
#         if np.max(self.Inputs.nStrands_inGroup)==1:
#             print('You are about to set a helium-fraction for a single-stranded wire!')
#
#         if not isinstance(self.Inputs.wBare_inGroup, np.ndarray):
#             self.Inputs.wBare_inGroup = np.array(self.Inputs.wBare_inGroup)
#         if not isinstance(self.Inputs.hBare_inGroup, np.ndarray):
#             self.Inputs.hBare_inGroup = np.array(self.Inputs.hBare_inGroup)
#         if not isinstance(self.Inputs.wIns_inGroup, np.ndarray):
#             self.Inputs.wIns_inGroup = np.array(self.Inputs.wIns_inGroup)
#         if not isinstance(self.Inputs.hIns_inGroup, np.ndarray):
#             self.Inputs.hIns_inGroup = np.array(self.Inputs.hIns_inGroup)
#         if not isinstance(self.Inputs.ds_inGroup, np.ndarray):
#             self.Inputs.ds_inGroup = np.array(self.Inputs.ds_inGroup)
#         if not isinstance(self.Inputs.nStrands_inGroup, np.ndarray):
#             self.Inputs.nStrands_inGroup = np.array(self.Inputs.nStrands_inGroup)
#
#         cs_bare = self.Inputs.wBare_inGroup*self.Inputs.hBare_inGroup
#         cs_ins = (self.Inputs.wBare_inGroup +2*self.Inputs.wIns_inGroup)* \
#                 (self.Inputs.hBare_inGroup +2*self.Inputs.hIns_inGroup)
#         cs_strand = self.Inputs.nStrands_inGroup*np.pi*(self.Inputs.ds_inGroup**2)/4
#         strand_total = cs_strand/cs_ins
#         ins_total = (cs_ins - cs_bare)/cs_ins
#         VoidRatio = (cs_bare - cs_strand)/cs_ins
#         extVoids = VoidRatio - (PercentVoids/100.0)
#         if any(sV < 0 for sV in extVoids):
#             print("Negative externalVoids calculated. Abort, please check.")
#             return
#         nGroups = len(self.Inputs.nT)
#         self.Inputs.overwrite_f_externalVoids_inGroup = extVoids
#         self.Inputs.overwrite_f_internalVoids_inGroup = np.ones((nGroups,)).transpose()*(PercentVoids/100.0)
#
#         self.descriptionsInputs['overwrite_f_externalVoids_inGroup'] = 'Helium fraction in the external cable voids'
#         self.descriptionsInputs['overwrite_f_internalVoids_inGroup'] = 'Helium fraction in the internal cable voids'
#
#     def preparePersistentCurrents(self, I_PC_LUT, dIdt, timeStep):
#         # LUT controlling power supply, Current [A]. Two cycles of ramping from 0 to nominal current and back to zero
#         if isinstance(I_PC_LUT,list):
#             I_PC_LUT = np.array(I_PC_LUT)
#         self.Inputs.I_PC_LUT = I_PC_LUT
#         self.Inputs.I00 = 0
#
#         # LUT controlling power supply, Time [s]
#         t_PC_LUT = np.zeros(len(self.Inputs.I_PC_LUT))
#         # Generates a time LUT that is dependent on the ramp rate of the current.
#         for x in range(len(self.Inputs.I_PC_LUT)):
#             if x == 0:  t_PC_LUT[x] = 0
#             elif x == 1: t_PC_LUT[x] = 0.1
#             elif x % 2 == 1: t_PC_LUT[x] = t_PC_LUT[x - 1] + 1
#             elif x % 4 == 0:
#                 t_PC_LUT[x] = t_PC_LUT[x - 1] - (self.Inputs.I_PC_LUT[x] - self.Inputs.I_PC_LUT[x - 1]) / dIdt
#             elif (x + 2) % 4 == 0:
#                 t_PC_LUT[x] = t_PC_LUT[x - 1] + (self.Inputs.I_PC_LUT[x] - self.Inputs.I_PC_LUT[x - 1]) / dIdt
#             else: continue
#         self.Inputs.t_PC_LUT =  t_PC_LUT
#
#         # time vector - Generates a time vector with finer timestepping when the ramp rate of the current changes
#         nElements = (len(self.Inputs.I_PC_LUT)-2)*6+3
#         time_vector_params = np.zeros(nElements)
#         every_sixth_element = range(nElements-3)[::6]
#         for x in every_sixth_element:
#             time_vector_params[x] = time_vector_params[x - 1] + timeStep
#             time_vector_params[x + 1] = timeStep
#             time_vector_params[x + 2] = t_PC_LUT[(x // 6) + 1] - 0.02
#             time_vector_params[x + 3] = time_vector_params[x + 2] + 0.001
#             time_vector_params[x + 4] = 0.001
#             time_vector_params[x + 5] = time_vector_params[x + 2] + 0.04
#         time_vector_params[0] = 0
#         time_vector_params[1] = 0.010
#         time_vector_params[-1] = t_PC_LUT[-1]
#         time_vector_params[-2] = timeStep
#         time_vector_params[-3] = time_vector_params[-4]+timeStep
#         self.Options.time_vector_params = time_vector_params
#
#         # Changes in options
#         if np.all(self.Inputs.f_SC_strand_inGroup == self.Inputs.f_SC_strand_inGroup[0]):
#             self.Options.flag_hotSpotTemperatureInEachGroup = 0
#         else:
#             self.Options.flag_hotSpotTemperatureInEachGroup = 0
#         self.Options.minCurrentDiode = 0
#         self.Options.flag_persistentCurrents = 1
#
#         # Changes in input
#         self.Inputs.t_PC = 99999
#         self.Inputs.tQH = np.array([99999]*len(self.Inputs.tQH))
#         self.Inputs.tEE = 99999
#         self.Inputs.tQuench = np.array([t_PC_LUT[-2]]*len(self.Inputs.M_m))
#         self.Inputs.tStartQuench = np.array([99999]*len(self.Inputs.tStartQuench))
#
#         selectedFont = {'fontname': 'DejaVu Sans', 'size': 14}
#         plt.figure(figsize=(5, 5))
#         plt.plot(self.Inputs.t_PC_LUT, self.Inputs.I_PC_LUT, 'ro-', label='LUT')
#         plt.xlabel('Time [s]', **selectedFont)
#         plt.ylabel('Current [A]', **selectedFont)
#         plt.title('Look-up table controlling power supply', **selectedFont)
#         plt.grid(True)
#         plt.rcParams.update({'font.size': 12})
#         plt.show()
#
#
#     def printVariableDescNameValue(self, variableGroup, variableLabels):
#         """
#
#            **Print variable description, variable name, and variable value**
#
#            Function prints variable description, variable name, and variable value
#
#            :param variableGroup: Dataclass containing all the attributes of the LEDET object
#            [obsolete, but still supported: list of tuples; each tuple has two elements: the first element is a string defining
#            the variable name, and the second element is either an integer, a float, a list, or a numpy.ndarray
#            defining the variable value :type variableGroup: list :param variableLabels: dictionary assigning a
#            description to each variable name]
#            :type variableLabels: dataclass [obsolete, but still supported: dict]
#
#            :return: None
#
#            [Example for usage of obsolete dictionary-version]
#             import numpy as np
#             variableGroup = []
#             variableGroup.append( ('x1', 12) )
#             variableGroup.append( ('x2', 23.42) )
#             variableGroup.append( ('x3', [2, 4, 6]) )
#             variableGroup.append( ('x3', np.array([2, 4, 6])) )
#             variableLabels = {'x1': '1st variable', 'x2': '2nd variable', 'x3': '3rd variable'}
#             utils.printVariableDescNameValue(variableGroup, variableLabels)
#             # >>> 					1st variable x1 12
#             # >>> 					2nd variable x2 23.42
#             # >>> 					3rd variable x3 [2, 4, 6]
#             # >>> 					3rd variable x3 [2 4 6]
#
#         """
#         if(variableGroup == asdict(self.Inputs)):
#             variableGroup = self.Inputs
#         if (variableGroup == asdict(self.Options)):
#             variableGroup = self.Options
#         if (variableGroup == asdict(self.Plots)):
#             variableGroup = self.Plots
#         if (variableGroup == asdict(self.Variables)):
#             variableGroup = self.Variables
#
#         if(type(variableGroup) != dict):
#             for k in variableGroup.__annotations__:
#                 if k == 'overwrite_f_internalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
#                 if k == 'overwrite_f_externalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
#                 print(variableLabels[k])
#                 print(k, self.getAttribute(variableGroup, k))
#         else:
#             for k in variableGroup:
#                 if k == 'overwrite_f_internalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
#                 if k == 'overwrite_f_externalVoids_inGroup' and len(self.getAttribute(variableGroup, k))==0: continue
#                 print(variableLabels[k])
#                 print(k, variableGroup[k])
